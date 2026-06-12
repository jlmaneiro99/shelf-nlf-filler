#!/usr/bin/env python3
"""Live verification of Railway /fill deployment. Never prints secrets."""
import base64
import io
import json
import os
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path
from typing import Optional, Tuple, Union

try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).parent / '.env')
    load_dotenv(Path(__file__).parent.parent / '.env')
except ImportError:
    pass

import openpyxl

DEFAULT_RAILWAY_URL = 'https://shelf-nlf-filler-production.up.railway.app'


def get_railway_url() -> str:
    url = (
        os.environ.get('RAILWAY_FILLER_URL')
        or os.environ.get('VITE_NLF_FILLER_URL')
        or DEFAULT_RAILWAY_URL
    )
    return url.rstrip('/')


def http_get(url: str) -> Tuple[int, Union[dict, str]]:
    req = urllib.request.Request(url, method='GET')
    with urllib.request.urlopen(req, timeout=60) as resp:
        body = resp.read().decode()
        try:
            return resp.status, json.loads(body)
        except json.JSONDecodeError:
            return resp.status, body


def http_post(url: str, payload: dict) -> Tuple[int, Union[dict, str]]:
    data = json.dumps(payload).encode()
    req = urllib.request.Request(
        url,
        data=data,
        method='POST',
        headers={'Content-Type': 'application/json'},
    )
    try:
        with urllib.request.urlopen(req, timeout=120) as resp:
            body = resp.read().decode()
            return resp.status, json.loads(body)
    except urllib.error.HTTPError as e:
        body = e.read().decode()
        try:
            return e.code, json.loads(body)
        except json.JSONDecodeError:
            return e.code, {'detail': body}


def build_test_workbook() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'LiveTest'
    ws['A1'] = 10
    ws['A4'] = 'Product Name'
    ws['A5'] = 'Brand Name'
    ws['A6'] = 'THR Licensed'
    ws['C4'] = '=A1*2'
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


PRODUCT = {
    'product_name': 'Live Test Product',
    'brand_name': 'VerifyBrand',
    'allergen_details': [],
    'certifications': [],
}


def run_fill(base_url: str) -> Tuple[int, Union[dict, str], Optional[dict]]:
    b64 = base64.b64encode(build_test_workbook()).decode()
    payload = {
        'file_base64': b64,
        'products': [PRODUCT],
        'retailer_name': 'LiveVerify',
        'fill_mode': 'auto',
        'form_spec': {
            'data_sheet': 'LiveTest',
            'layout': 'vertical',
            'label_column': 1,
            'value_column': 2,
            'example_rows': [],
            'other_sheets': [],
            'field_map': [],
        },
    }
    status, body = http_post(f'{base_url}/fill', payload)
    if status != 200 or not isinstance(body, dict) or 'file_base64' not in body:
        return status, body, None
    out = openpyxl.load_workbook(io.BytesIO(base64.b64decode(body['file_base64'])))
    ws = out['LiveTest']
    checks = {
        'product_name': ws['B4'].value,
        'brand_name': ws['B5'].value,
        'thr_licensed': ws['B6'].value,
        'formula_c4': ws['C4'].value,
        'fields_filled': body.get('fields_filled'),
        'anthropic_key_present': body.get('anthropic_key_present'),
    }
    return status, body, checks


def wait_for_config_endpoint(base_url: str, max_wait: int = 180) -> bool:
    """Poll until /health/config exists (post-deploy)."""
    deadline = time.time() + max_wait
    while time.time() < deadline:
        try:
            status, body = http_get(f'{base_url}/health/config')
            if status == 200 and isinstance(body, dict):
                return True
        except urllib.error.HTTPError as e:
            if e.code != 404:
                raise
        except Exception:
            pass
        time.sleep(5)
    return False


def pf(ok: bool) -> str:
    return 'PASS' if ok else 'FAIL'


def main():
    base = get_railway_url()
    host = urllib.parse.urlparse(base).netloc
    results = []

    print(f'Railway host: {host}')
    print()

    print('--- GET /health ---')
    try:
        status, body = http_get(f'{base}/health')
        health_ok = status == 200 and isinstance(body, dict) and body.get('status') == 'ok'
        print(f'{pf(health_ok)} | /health status={status} body={body}')
        results.append(health_ok)
    except Exception as e:
        print(f'FAIL | /health error: {type(e).__name__}: {e}')
        results.append(False)
    print()

    print('--- GET /health/config ---')
    config_ok = False
    key_present = False
    try:
        status, body = http_get(f'{base}/health/config')
    except urllib.error.HTTPError as e:
        if e.code == 404:
            print('WARN | /health/config not found yet — waiting for deploy...')
            if wait_for_config_endpoint(base):
                status, body = http_get(f'{base}/health/config')
            else:
                status, body = 404, {}
        else:
            raise

    if status == 200 and isinstance(body, dict):
        key_present = bool(body.get('anthropic_key_present'))
        httpx_ok = bool(body.get('anthropic_client_available'))
        config_ok = key_present and httpx_ok
        print(f'{pf(config_ok)} | anthropic_key_present={body.get("anthropic_key_present")} '
              f'source={body.get("anthropic_key_source")!r} '
              f'client_available={body.get("anthropic_client_available")}')
    else:
        print(f'FAIL | /health/config status={status}')
    results.append(config_ok)
    print()

    print('--- POST /fill (live) ---')
    status, body, checks = run_fill(base)
    if checks:
        known_ok = (
            checks.get('product_name') == 'Live Test Product'
            and checks.get('brand_name') == 'VerifyBrand'
        )
        formula_ok = (
            isinstance(checks.get('formula_c4'), str)
            and str(checks.get('formula_c4')).startswith('=')
        )
        thr = checks.get('thr_licensed')
        anthropic_ok = thr is not None and str(thr).strip() != ''
        fill_key_flag = checks.get('anthropic_key_present')

        print(f'{pf(status == 200)} | HTTP status={status}')
        print(f'{pf(known_ok)} | known fields B4={checks.get("product_name")!r} B5={checks.get("brand_name")!r}')
        print(f'{pf(formula_ok)} | formula C4={checks.get("formula_c4")!r}')
        print(f'{pf(anthropic_ok)} | THR Licensed B6={checks.get("thr_licensed")!r} fields_filled={checks.get("fields_filled")}')
        print(f'{pf(fill_key_flag is True)} | fill response anthropic_key_present={fill_key_flag}')

        results.extend([status == 200, known_ok, formula_ok, anthropic_ok, fill_key_flag is True])
    else:
        print(f'FAIL | /fill status={status} body={str(body)[:300]}')
        results.extend([False, False, False, False, False])
    print()

    passed = sum(1 for r in results if r)
    total = len(results)
    all_pass = passed == total
    print(f'SUMMARY: {passed}/{total} checks passed')
    if all_pass:
        print('VERDICT: PASS — READY for Dundeis real-form test')
        return 0
    print('VERDICT: FAIL — see checks above')
    return 1


if __name__ == '__main__':
    sys.exit(main())
