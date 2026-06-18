import json
import asyncio
import base64
import io

import openpyxl
from openpyxl.cell.cell import MergedCell

from main import (
    map_field,
    safe_str,
    is_allergen_field,
    safe_write,
    is_formula_cell,
    count_formula_cells,
    fill_horizontal_rows,
    fill_using_form_spec,
    FillRequest,
    fill_nlf,
)

TEST_PRODUCT = {
    'product_name': 'Test Protein Powder',
    'brand_name': 'TestBrand',
    'supplier_name': 'TestBrand Ltd',
    'sku_code': 'TB-001',
    'ean_barcode': '5060000000001',
    'case_barcode': None,
    'variant': '500g',
    'rrp': 19.99,
    'trade_price_per_case': 95.00,
    'units_per_case': 6,
    'case_size_description': '6 x 500g',
    'vat_rate': 'Zero',
    'tax_status': 'Zero-rated',
    'shelf_life_weeks': 52,
    'storage_conditions': 'Ambient',
    'country_of_origin': 'United Kingdom',
    'is_vegan': True,
    'is_vegetarian': True,
    'is_gluten_free': True,
    'is_organic': False,
    'organic_cert_number': None,
    'is_recyclable': True,
    'contains_added_sugar': False,
    'energy_kcal': 400,
    'energy_kj': 1674,
    'fat': 5.0,
    'saturates': 1.0,
    'carbohydrates': 10.0,
    'sugars': 2.0,
    'fibre': 3.0,
    'protein': 75.0,
    'salt': 1.0,
    'serving_size_value': 30,
    'allergen_details': [
        {'allergen': 'Milk', 'may_contain': True}
    ],
    'lead_time_days': None,
    'certifications': ['Vegan', 'Gluten Free'],
}


def test(label, expected, product=TEST_PRODUCT):
    result = map_field(label, product)
    status = 'PASS' if str(result) == str(expected) else 'FAIL'
    print(f'{status} | "{label}" → got "{result}" expected "{expected}"')
    return status == 'PASS'


results = []

results.append(test('Full Product Name *', 'Test Protein Powder'))
results.append(test('Product Name on Pack', 'Test Protein Powder'))
results.append(test('Brand Name', 'TestBrand'))
results.append(test('Supplier Name', 'TestBrand Ltd'))
results.append(test('SKU / Supplier Code', 'TB-001'))
results.append(test('EAN Barcode (Unit)', '5060000000001'))
results.append(test('Individual unit barcode', '5060000000001'))
results.append(test('Case Barcode', 'N/A'))
results.append(test('UPC Barcode Case', 'N/A'))
results.append(test('Item Description', 'Test Protein Powder'))
results.append(test('Item Name', 'Test Protein Powder'))
results.append(test('UPC Barcode', '5060000000001'))

case_upc_product = {**TEST_PRODUCT, 'ean_barcode': '5060000000001', 'case_barcode': '19310000444009'}
results.append(test('UPC Barcode Case', '19310000444009', product=case_upc_product))
results.append(test('UPC Barcode', '5060000000001', product=case_upc_product))

results.append(test('RRP per Unit (inc VAT)', '19.99'))
results.append(test('MSRP per Unit (USD)', '19.99'))
results.append(test('Wholesale Price per Case (ex VAT)', '95.0'))
results.append(test('Trade Price per Case', '95.0'))
results.append(test('Units per Case', '6'))
results.append(test('VAT Rate', 'Zero-rated'))
results.append(test('Lead Time (days)', 'N/A'))

results.append(test('Shelf Life (weeks from manufacture)', '52'))
results.append(test('Shelf Life from Manufacture days', '364'))
results.append(test('Storage Conditions', 'Ambient'))
results.append(test('Country of Origin', 'United Kingdom'))

results.append(test('Is product Vegan?', 'Yes'))
results.append(test('Vegan', 'Yes'))
results.append(test('Is product Gluten Free?', 'Yes'))
results.append(test('Gluten Free', 'Yes'))
results.append(test('Gluten Free?', 'Yes'))
results.append(test('Is product Organic?', 'No'))
results.append(test('Organic Certification Number', 'N/A'))
results.append(test('Contains Added Sugar?', 'No'))
results.append(test('Is Packaging Recyclable', 'Yes'))
results.append(test('Is packaging recyclable?', 'Yes'))

results.append(test('Energy kcal per 100g', '400'))
results.append(test('Fat g per 100g', '5'))
results.append(test('Protein g per 100g', '75'))
results.append(test('Salt g per 100g', '1'))

results.append(test('Total Fat g per serving', '1.5'))
results.append(test('Total Protein g per serving', '22.5'))

results.append(test('Sodium mg per 100g', '400'))
results.append(test('Sodium mg per serving', '120'))

results.append(test('Added Sugars g per serving', '0'))

results.append(test('Milk', 'May contain'))
results.append(test('Milk / Dairy', 'May contain'))
results.append(test('Eggs', 'Not present'))
results.append(test('Peanuts', 'Not present'))
results.append(test('Cereals Containing Gluten', 'Not present'))

hfss_product = {**TEST_PRODUCT, 'hfss_score': 4}
results.append(test('HFSS Score (Nutrient Profiling Score)', '4', product=hfss_product))

vegsoc_product = {**TEST_PRODUCT, 'vegsoc_trademark': False, 'ingredients': 'must not leak here'}
results.append(test('Vegetarian society trademark criteria', 'No', product=vegsoc_product))

# Conservative compliance — NASAA must not infer from generic Organic
nasaa_product = {**TEST_PRODUCT, 'is_organic': True, 'certifications': ['Organic', 'Vegan']}
results.append(test('NASAA Organic', 'No', product=nasaa_product))
results.append(test('Australian Certified Organic', 'No', product=nasaa_product))

# Unknown label returns None (Claude fallback candidate)
thr = map_field('THR Licensed', TEST_PRODUCT)
thr_ok = thr is None
print(f'{"PASS" if thr_ok else "FAIL"} | "THR Licensed" → got "{thr}" expected None')
results.append(thr_ok)

print()
print('--- Integration tests ---')

# Formula protection — value cell is empty; unrelated formula preserved
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Data'
ws['A1'] = 5
ws['A4'] = 'Product Name'
ws['C4'] = '=A1*2'
file_bytes = io.BytesIO()
wb.save(file_bytes)
b64 = base64.b64encode(file_bytes.getvalue()).decode()

req = FillRequest(
    file_base64=b64,
    products=[{'product_name': 'Test Item', 'allergen_details': [], 'certifications': []}],
    retailer_name='Test',
    fill_mode='auto',
    form_spec={
        'data_sheet': 'Data',
        'layout': 'vertical',
        'label_column': 1,
        'value_column': 2,
        'example_rows': [],
        'other_sheets': [],
        'field_map': [],
    },
)
res = asyncio.run(fill_nlf(req))
out_wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(res['file_base64'])))
formula_val = out_wb['Data']['C4'].value
value_filled = out_wb['Data']['B4'].value == 'Test Item'
formula_ok = isinstance(formula_val, str) and formula_val.startswith('=') and value_filled
print(f'{"PASS" if formula_ok else "FAIL"} | Formula preserved + value filled: formula={formula_val!r} value={out_wb["Data"]["B4"].value!r}')
results.append(formula_ok)

# Formula in value column → skip silently, file still returned, formula intact
wb_formula_target = openpyxl.Workbook()
ws_ft = wb_formula_target.active
ws_ft.title = 'Data'
ws_ft['A1'] = 5
ws_ft['A4'] = 'Product Name'
ws_ft['B4'] = '=A1*2'
ws_ft['A5'] = 'Brand'
file_bytes_ft = io.BytesIO()
wb_formula_target.save(file_bytes_ft)
b64_ft = base64.b64encode(file_bytes_ft.getvalue()).decode()
req_ft = FillRequest(
    file_base64=b64_ft,
    products=[{
        'product_name': 'Should Not Overwrite',
        'brand_name': 'TestBrand',
        'allergen_details': [],
        'certifications': [],
    }],
    retailer_name='Test',
    form_spec={
        'data_sheet': 'Data',
        'layout': 'vertical',
        'label_column': 1,
        'value_column': 2,
        'example_rows': [],
        'other_sheets': [],
        'field_map': [],
    },
)
res_ft = asyncio.run(fill_nlf(req_ft))
out_ft = openpyxl.load_workbook(io.BytesIO(base64.b64decode(res_ft['file_base64'])))
formula_skip_ok = (
    isinstance(out_ft['Data']['B4'].value, str)
    and str(out_ft['Data']['B4'].value).startswith('=')
    and out_ft['Data']['B5'].value == 'TestBrand'
    and res_ft.get('fields_filled', 0) >= 1
)
print(f'{"PASS" if formula_skip_ok else "FAIL"} | Formula value cell skipped, other cells filled, file returned')
results.append(formula_skip_ok)

# Example row protection
wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = 'Sheet1'
for col, hdr in enumerate(['Product Name', 'Brand', 'RRP'], start=1):
    ws2.cell(row=8, column=col).value = hdr
ws2.cell(row=9, column=1).value = 'EXAMPLE PRODUCT'
ws2.cell(row=9, column=2).value = 'Demo Brand'
ws2.cell(row=9, column=3).value = '9.99'
file_bytes2 = io.BytesIO()
wb2.save(file_bytes2)
b64_2 = base64.b64encode(file_bytes2.getvalue()).decode()

products_3 = [
    {'product_name': f'Product {i}', 'brand_name': 'BrandX', 'rrp': 10 + i,
     'allergen_details': [], 'certifications': []}
    for i in range(1, 4)
]
req2 = FillRequest(
    file_base64=b64_2,
    products=products_3,
    retailer_name='Dundeis',
    fill_mode='auto',
    form_spec={
        'data_sheet': 'Sheet1',
        'layout': 'horizontal_rows',
        'header_row': 8,
        'first_data_row': 10,
        'example_rows': [9],
        'other_sheets': [],
        'field_map': [],
    },
)
res2 = asyncio.run(fill_nlf(req2))
out2 = openpyxl.load_workbook(io.BytesIO(base64.b64decode(res2['file_base64'])))
ws_out = out2['Sheet1']
example_preserved = ws_out.cell(row=9, column=1).value == 'EXAMPLE PRODUCT'
row10 = ws_out.cell(row=10, column=1).value == 'Product 1'
row11 = ws_out.cell(row=11, column=1).value == 'Product 2'
row12 = ws_out.cell(row=12, column=1).value == 'Product 3'
example_ok = example_preserved and row10 and row11 and row12
print(f'{"PASS" if example_ok else "FAIL"} | Example row untouched + products at 10/11/12')
results.append(example_ok)

# Formula count unchanged with VLOOKUP sheet
wb3 = openpyxl.Workbook()
ws_data = wb3.active
ws_data.title = 'Product 1'
ws_data['A4'] = 'Product Name'
ws_data['B4'] = 'Brand'
ws_data['A5'] = 'Full Product Name'
ws_lookup = wb3.create_sheet('Lookup')
ws_lookup['A1'] = '=VLOOKUP("x",Product_1!A:B,2,FALSE)'
file_bytes3 = io.BytesIO()
wb3.save(file_bytes3)
before_count = count_formula_cells(openpyxl.load_workbook(io.BytesIO(file_bytes3.getvalue())))
b64_3 = base64.b64encode(file_bytes3.getvalue()).decode()
req3 = FillRequest(
    file_base64=b64_3,
    products=[{'product_name': 'Safe Fill', 'brand_name': 'Co', 'allergen_details': [], 'certifications': []}],
    retailer_name='Test',
    form_spec={
        'data_sheet': 'Product 1',
        'layout': 'vertical',
        'label_column': 1,
        'value_column': 2,
        'other_sheets': ['Lookup'],
        'example_rows': [],
        'field_map': [],
    },
)
res3 = asyncio.run(fill_nlf(req3))
after_wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(res3['file_base64'])))
after_count = count_formula_cells(after_wb)
count_ok = after_count >= before_count and 'Lookup' in after_wb.sheetnames
print(f'{"PASS" if count_ok else "FAIL"} | Formula count before={before_count} after={after_count}')
results.append(count_ok)

# Dundeis-style: 5 products, horizontal_rows, example row, Instructions sheet untouched
# Sheet name has trailing space — form_spec uses name without space (must still resolve)
wb5 = openpyxl.Workbook()
ws5 = wb5.active
ws5.title = 'Product details '
headers5 = ['Product Name', 'Brand', 'RRP', 'EAN Barcode (Unit)', 'Vegan']
for col, hdr in enumerate(headers5, start=2):
    ws5.cell(row=8, column=col).value = hdr
ws5.cell(row=9, column=2).value = '1. Example line'
ws5.cell(row=9, column=3).value = 'Salted Peanuts'
instr = wb5.create_sheet('Marketing content')
instr['A1'] = 'Do not modify this sheet'
instr['B2'] = '=1+1'
file_bytes5 = io.BytesIO()
wb5.save(file_bytes5)
b64_5 = base64.b64encode(file_bytes5.getvalue()).decode()
products_5 = [
    {'product_name': f'Product {i}', 'brand_name': 'Mi-Eco', 'rrp': 9.99 + i,
     'ean_barcode': f'506000000000{i}', 'is_vegan': True,
     'allergen_details': [], 'certifications': []}
    for i in range(1, 6)
]
req5 = FillRequest(
    file_base64=b64_5,
    products=products_5,
    retailer_name='Retailer_NLF_5_products',
    fill_mode='auto',
    form_spec={
        'data_sheet': 'Product details',
        'layout': 'horizontal_rows',
        'header_row': 8,
        'first_data_row': 10,
        'example_rows': [9],
        'other_sheets': ['Marketing content'],
        'field_map': [],
    },
)
res5 = asyncio.run(fill_nlf(req5))
out5 = openpyxl.load_workbook(io.BytesIO(base64.b64decode(res5['file_base64'])))
ws5_out = out5['Product details ']
dundeis_ok = (
    res5['fields_filled'] > 0
    and ws5_out.cell(row=9, column=2).value == '1. Example line'
    and ws5_out.cell(row=9, column=3).value == 'Salted Peanuts'
    and ws5_out.cell(row=10, column=2).value == 'Product 1'
    and ws5_out.cell(row=14, column=2).value == 'Product 5'
    and ws5_out.cell(row=10, column=5).value == '5060000000001'
    and ws5_out.cell(row=10, column=6).value == 'Yes'
    and out5['Marketing content']['A1'].value == 'Do not modify this sheet'
    and str(out5['Marketing content']['B2'].value).startswith('=')
)
print(f'{"PASS" if dundeis_ok else "FAIL"} | Dundeis 5-product horizontal_rows + trailing-space sheet + example preserved')
results.append(dundeis_ok)

# Trailing-space sheet name resolves via resolve_sheet_name
from main import resolve_sheet_name
trail_wb = openpyxl.Workbook()
trail_wb.active.title = 'Product details '
trail_ok = resolve_sheet_name(trail_wb, 'Product details') == 'Product details '
print(f'{"PASS" if trail_ok else "FAIL"} | Trailing-space sheet name resolves correctly')
results.append(trail_ok)

# Zero-fill must raise 422, never return blank file
from fastapi import HTTPException
wb_zero = openpyxl.Workbook()
wb_zero.active.title = 'Sheet1'
zero_buf = io.BytesIO()
wb_zero.save(zero_buf)
b64_zero = base64.b64encode(zero_buf.getvalue()).decode()
req_zero = FillRequest(
    file_base64=b64_zero,
    products=[{'product_name': 'X', 'brand_name': 'Y', 'allergen_details': [], 'certifications': []}],
    retailer_name='Test',
    fill_mode='auto',
    form_spec={
        'data_sheet': 'Sheet1',
        'layout': 'vertical',
        'label_column': 99,
        'value_column': 100,
        'other_sheets': [],
        'example_rows': [],
        'field_map': [],
    },
)
zero_fill_ok = False
try:
    asyncio.run(fill_nlf(req_zero))
except HTTPException as exc:
    zero_fill_ok = exc.status_code == 422 and 'No fields were filled' in str(exc.detail)
print(f'{"PASS" if zero_fill_ok else "FAIL"} | Zero-fill raises 422 with diagnostic detail')
results.append(zero_fill_ok)

# Precomputed mappings write Step 3 values even when map_field would fail
wb_pre = openpyxl.Workbook()
ws_pre = wb_pre.active
ws_pre.title = 'Product details '
for col, hdr in enumerate(['Product Name', 'Brand', 'RRP'], start=2):
    ws_pre.cell(row=8, column=col).value = hdr
ws_pre.cell(row=9, column=2).value = 'Example line'
pre_buf = io.BytesIO()
wb_pre.save(pre_buf)
b64_pre = base64.b64encode(pre_buf.getvalue()).decode()
req_pre = FillRequest(
    file_base64=b64_pre,
    products=[{'product_name': 'IGNORED', 'brand_name': 'X', 'allergen_details': [], 'certifications': []}],
    retailer_name='Precomputed',
    fill_mode='auto',
    form_spec={
        'data_sheet': 'Product details',
        'layout': 'horizontal_rows',
        'header_row': 8,
        'first_data_row': 10,
        'example_rows': [9],
        'other_sheets': [],
        'field_map': [],
    },
    precomputed_mappings=[
        {'sheet_name': 'Product details', 'row': 10, 'col': 2, 'value': 'From Step 3', 'field_label': 'Product Name'},
        {'sheet_name': 'Product details', 'row': 10, 'col': 3, 'value': 'Brand Co', 'field_label': 'Brand'},
        {'sheet_name': 'Product details', 'row': 10, 'col': 4, 'value': '12.99', 'field_label': 'RRP'},
    ],
)
res_pre = asyncio.run(fill_nlf(req_pre))
out_pre = openpyxl.load_workbook(io.BytesIO(base64.b64decode(res_pre['file_base64'])))
ws_pre_out = out_pre['Product details ']
precomputed_ok = (
    res_pre['fields_filled'] > 0
    and ws_pre_out.cell(row=10, column=2).value == 'From Step 3'
    and ws_pre_out.cell(row=9, column=2).value == 'Example line'
)
print(f'{"PASS" if precomputed_ok else "FAIL"} | Precomputed Step 3 mappings written to sheet')
results.append(precomputed_ok)

print()
print('--- Double-layout / single-pass conflict tests ---')

import main as _m
from openpyxl.utils import get_column_letter as _gcl


def _build_dundeis_wb():
    """73-ish label horizontal form: header row 8, example row 9, formula col V,
    a junk numeric header, a commodity column, and bare packaging headers."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Product details '  # trailing space on purpose
    labels = {
        2: 'Product Name', 3: 'Brand', 4: 'RRP', 5: 'Wholesale Price per Case (ex VAT)',
        6: 'Storage Conditions', 7: 'Country of Origin', 8: 'Commodity Code (10 digit)',
        9: 'Other', 10: 'Paper', 11: 'Energy kcal per 100g', 12: 'Protein g per 100g',
    }
    for col, text in labels.items():
        ws.cell(row=8, column=col).value = text
    ws.cell(row=8, column=22).value = 'RRP Margin'   # col V = formula field
    ws.cell(row=8, column=23).value = '1.65'          # junk numeric header (col W)
    # example row 9
    ws.cell(row=9, column=2).value = '1. Example line'
    ws.cell(row=9, column=3).value = 'Salted Peanuts'
    ws.cell(row=9, column=9).value = '13g'            # example junk for "Other"
    ws.cell(row=9, column=23).value = 'EXAMPLE'
    # formula in col V for the (future) product rows so it must be protected
    for r in range(10, 15):
        ws.cell(row=r, column=22).value = '=(U{0}-(S{0}*(1+W{0})))/U{0}'.format(r)
    other = wb.create_sheet('Marketing content')
    other['A1'] = 'do not touch'
    other['B2'] = '=A1'
    return wb


dundeis_products = [
    {
        'product_name': f'Granola {i}', 'brand_name': 'Mi-Eco', 'rrp': 4.99 + i,
        'trade_price_per_case': 17.52, 'storage_conditions': 'Ambient',
        'country_of_origin': 'Bulgaria', 'sku_code': 'FN-PEA-VAN-500',
        'hs_commodity_code': '1904100000', 'energy_kcal': 422, 'protein': 17,
        'allergen_details': [], 'certifications': [],
    }
    for i in range(5)
]
spec_h = {
    'data_sheet': 'Product details',
    'layout': 'horizontal_rows',
    'header_row': 8,
    'first_data_row': 10,
    'example_rows': [9],
    'other_sheets': ['Marketing content'],
    'field_map': [],
}


def _run_dundeis(precomputed=None):
    wb = _build_dundeis_wb()
    buf = io.BytesIO()
    wb.save(buf)
    req = FillRequest(
        file_base64=base64.b64encode(buf.getvalue()).decode(),
        products=dundeis_products,
        retailer_name='Dundeis',
        fill_mode='auto',
        form_spec=spec_h,
        precomputed_mappings=precomputed,
    )
    res = asyncio.run(fill_nlf(req))
    out = openpyxl.load_workbook(io.BytesIO(base64.b64decode(res['file_base64'])))
    return res, out['Product details ']


# 5 products in rows 10-14, NO vertical dump in any single column
res_d, ws_d = _run_dundeis()
rows_ok = all(ws_d.cell(row=10 + i, column=2).value == f'Granola {i}' for i in range(5))
# column C (3) must NOT contain a stacked vertical product (only header 8 +
# example row 9 + product rows 10-14 may be populated)
col_c_rows = [r for r in range(9, 90) if ws_d.cell(row=r, column=3).value not in (None, '')]
no_vertical_dump = set(col_c_rows).issubset({9, 10, 11, 12, 13, 14})  # 9 = example brand
example_safe = ws_d.cell(row=9, column=2).value == '1. Example line'
print(f'{"PASS" if rows_ok and no_vertical_dump and example_safe else "FAIL"} '
      f'| 5 products in rows 10-14, no vertical dump in column C, example row intact')
results.append(rows_ok and no_vertical_dump and example_safe)

# Formula column V never overwritten on any product row
v_formula_ok = all(
    str(ws_d.cell(row=10 + i, column=22).value).startswith('=') for i in range(5)
)
print(f'{"PASS" if v_formula_ok else "FAIL"} | Formula column V preserved on all product rows')
results.append(v_formula_ok)

# Junk numeric header column ("1.65", col W=23) skipped entirely
junk_skipped = all(ws_d.cell(row=10 + i, column=23).value in (None, '') for i in range(5))
print(f'{"PASS" if junk_skipped else "FAIL"} | Numeric/junk header column ("1.65") skipped')
results.append(junk_skipped)

# Commodity code column gets the HS code, NEVER the SKU
commodity_val = ws_d.cell(row=10, column=8).value
commodity_ok = commodity_val == '1904100000' and commodity_val != 'FN-PEA-VAN-500'
print(f'{"PASS" if commodity_ok else "FAIL"} | Commodity code = HS code, not SKU (got {commodity_val!r})')
results.append(commodity_ok)

# Bare "Other"/"Paper" packaging headers get N/A (never example junk like "13g")
packaging_blank = all(
    ws_d.cell(row=10 + i, column=9).value in (None, '', 'N/A')
    and ws_d.cell(row=10 + i, column=10).value in (None, '', 'N/A')
    for i in range(5)
)
print(f'{"PASS" if packaging_blank else "FAIL"} | Bare packaging headers ("Other"/"Paper") N/A, no example leak')
results.append(packaging_blank)

# Ancestral-style unit size, UOM, trade cost labels
_ancestral = {
    'product_name': 'Alchemy Bites',
    'unit_net_weight_g': 40,
    'trade_price_per_case': 14.5,
    'units_per_case': 12,
    'weight_unit': 'g',
    'cost_price_per_case': None,
    'inner_plastic_weight_g': 13,
}
_trade_unit = _m.trade_price_per_unit(_ancestral)
ancestral_fields_ok = (
    _m.map_field('Unit Size', _ancestral) == '40'
    and _m.map_field('UOM', _ancestral) == 'g'
    and _m.map_field('Trade Case Cost', _ancestral) == '14.5'
    and _m.map_field('Case Cost - Trade', _ancestral) == '14.5'
    and _m.map_field('Trade Unit Cost', _ancestral) == _trade_unit
    and _m.map_field('Dundeis Case Cost', _ancestral) == 'N/A'
    and _m.map_field('Dundeis Unit Cost', _ancestral) == 'N/A'
    and _m.map_field('Other', _ancestral) == 'N/A'
    and _m.map_field('Product Name & Description', _ancestral) == 'Alchemy Bites'
)
print(f'{"PASS" if ancestral_fields_ok else "FAIL"} | Unit Size/UOM/trade costs + Other=N/A for Ancestral shape')
results.append(ancestral_fields_ok)

# Variant-only fallback when unit_net_weight_g missing
_variant_only = {
    'variant': '40g',
    'trade_price_per_case': 14.5,
    'units_per_case': 12,
}
variant_fallback_ok = (
    _m.map_field('Unit Size', _variant_only) == '40'
    and _m.map_field('UOM', _variant_only) == 'g'
)
print(f'{"PASS" if variant_fallback_ok else "FAIL"} | Unit Size/UOM from variant when weight missing')
results.append(variant_fallback_ok)

# Empty precomputed must not block map_field
pre_empty_ok = True
try:
    wb_pe = _build_dundeis_wb()
    wb_pe.active.cell(row=8, column=24).value = 'Unit Size'
    wb_pe.active.cell(row=8, column=25).value = 'Trade Case Cost'
    spec_pe = {**spec_h, 'field_map': [
        {'label': 'Unit Size', 'col': 24},
        {'label': 'Trade Case Cost', 'col': 25},
    ]}
    plan_pe = _m.resolve_fill_plan(wb_pe, spec_pe, 'auto')
    _m._WriteTracker.reset()
    ws_pe = wb_pe[plan_pe['sheet_used']]
    ancestral_fill = [{
        'product_name': 'Alchemy Bites', 'unit_net_weight_g': 40,
        'trade_price_per_case': 14.5, 'units_per_case': 12, 'allergen_details': [],
    }]
    filled_pe, _, _ = _m.fill_horizontal_rows(
        ws_pe, ancestral_fill, plan_pe,
        precomputed=[{
            'field_label': 'Unit Size', 'value': '', 'product_index': 0,
        }, {
            'field_label': 'Trade Case Cost', 'value': '   ', 'product_index': 0,
        }],
    )
    pre_empty_ok = (
        ws_pe.cell(row=10, column=24).value == '40'
        and ws_pe.cell(row=10, column=25).value == '14.5'
    )
except Exception as exc:
    print(f'pre_empty test error: {exc}')
    pre_empty_ok = False
print(f'{"PASS" if pre_empty_ok else "FAIL"} | Empty precomputed falls back to map_field')
results.append(pre_empty_ok)

# map_field never returns SKU for commodity/hs/tariff/meursing
commodity_map_ok = (
    _m.map_field('Commodity Code (10 digit)', dundeis_products[0]) == '1904100000'
    and _m.map_field('HS Code', dundeis_products[0]) == '1904100000'
    and _m.map_field('Tariff Code', dundeis_products[0]) != 'FN-PEA-VAN-500'
)
print(f'{"PASS" if commodity_map_ok else "FAIL"} | map_field: commodity/hs/tariff never return SKU')
results.append(commodity_map_ok)

# Example row values never propagate: a field not in the vault stays blank
example_no_leak = _m.map_field('Shelf Ready Case', dundeis_products[0]) is None
print(f'{"PASS" if example_no_leak else "FAIL"} | Unknown vault field returns None (no example leakage)')
results.append(example_no_leak)

# A stray vertical-coordinate precomputed payload on a horizontal form is
# rebuilt to horizontal rows — it must NOT create a vertical column dump.
vertical_payload = [
    {'sheet_name': 'Product details', 'row': 4, 'col': 3, 'value': 'VDUMP-Name', 'field_label': 'Product Name', 'product_index': 0},
    {'sheet_name': 'Product details', 'row': 15, 'col': 3, 'value': 'Ambient', 'field_label': 'Storage Conditions', 'product_index': 0},
    {'sheet_name': 'Product details', 'row': 75, 'col': 3, 'value': '422', 'field_label': 'Energy kcal per 100g', 'product_index': 0},
]
res_v, ws_v = _run_dundeis(precomputed=vertical_payload)
# rows 4/15/75 in column C must be untouched; product name landed in row 10
c4 = ws_v.cell(row=4, column=3).value
c15 = ws_v.cell(row=15, column=3).value
c75 = ws_v.cell(row=75, column=3).value
no_legacy_vertical = c4 in (None, '') and c15 in (None, '') and c75 in (None, '')
print(f'{"PASS" if no_legacy_vertical else "FAIL"} '
      f'| Vertical-coord precomputed rebuilt horizontally (C4/C15/C75 untouched)')
results.append(no_legacy_vertical)

# Conflict guard: a genuine vertical dump injected via safe_write triggers 422
conflict_caught = False
try:
    wb_cf = _build_dundeis_wb()
    plan_cf = _m.resolve_fill_plan(wb_cf, spec_h, 'auto')
    _m._WriteTracker.reset()
    ws_cf = wb_cf[plan_cf['sheet_used']]
    # simulate horizontal fill (wide rows 10-14)
    for i in range(5):
        for c in range(2, 13):
            _m.safe_write(ws_cf, 10 + i, c, 'x')
    # simulate stray vertical dump down column C across many non-product rows
    for r in [4, 5, 6, 15, 51, 75, 76, 77]:
        _m.safe_write(ws_cf, r, 3, 'dump')
    conflict_msg = _m.check_write_conflict(plan_cf, 5)
    conflict_caught = conflict_msg is not None and 'conflict' in conflict_msg.lower()
except Exception:
    conflict_caught = False
print(f'{"PASS" if conflict_caught else "FAIL"} | Conflicting double-layout write triggers conflict (422)')
results.append(conflict_caught)

# Vertical writer must NOT run on a horizontal form — even with bait labels in
# column B at rows 15/75 and NO form_spec (detection-only). C15/C75 stay empty.
wb_bait = openpyxl.Workbook()
ws_bait = wb_bait.active
ws_bait.title = 'Product details '
_bait_headers = {
    2: 'Product Name', 3: 'Brand', 4: 'RRP', 5: 'Storage Conditions',
    6: 'Country of Origin', 7: 'Energy kcal per 100g', 8: 'Protein g per 100g',
}
for _c, _t in _bait_headers.items():
    ws_bait.cell(row=8, column=_c).value = _t
ws_bait.cell(row=9, column=2).value = '1. Example line'
# Bait: legacy Suma vertical label positions in column B
ws_bait.cell(row=15, column=2).value = 'Storage Conditions'
ws_bait.cell(row=75, column=2).value = 'Energy kcal per 100g'
_bait_buf = io.BytesIO()
wb_bait.save(_bait_buf)
bait_products = [
    {'product_name': f'Granola {i}', 'brand_name': 'Mi-Eco', 'rrp': 4.99 + i,
     'storage_conditions': 'Ambient', 'country_of_origin': 'Bulgaria',
     'energy_kcal': 422, 'protein': 17, 'allergen_details': [], 'certifications': []}
    for i in range(5)
]
# NO form_spec — force Railway's own layout detection
req_bait = FillRequest(
    file_base64=base64.b64encode(_bait_buf.getvalue()).decode(),
    products=bait_products,
    retailer_name='Dundeis',
    fill_mode='auto',
    form_spec=None,
)
res_bait = asyncio.run(fill_nlf(req_bait))
out_bait = openpyxl.load_workbook(io.BytesIO(base64.b64decode(res_bait['file_base64'])))
ws_bait_out = out_bait['Product details ']
c15 = ws_bait_out.cell(row=15, column=3).value
c75 = ws_bait_out.cell(row=75, column=3).value
horizontal_filled = ws_bait_out.cell(row=10, column=2).value == 'Granola 0'
no_vertical = c15 in (None, '') and c75 in (None, '')
print(f'{"PASS" if no_vertical and horizontal_filled else "FAIL"} '
      f'| Horizontal form (no form_spec): C15={c15!r} C75={c75!r} empty, rows filled')
results.append(no_vertical and horizontal_filled)

# Tabs mode: Step 3 precomputed for product 2 must NOT contaminate Product 1 tab
def _build_suma_tabs_wb():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Product 1'
    ws.cell(row=4, column=2).value = 'Suma Product Code'
    ws.cell(row=5, column=2).value = 'Full Product Name'
    ws.cell(row=6, column=2).value = 'Brand'
    ws.cell(row=7, column=2).value = 'Supplier reference code'
    # Stale template junk (simulates Cold Brew example left on template tab)
    ws.cell(row=5, column=3).value = 'Original Cold Brew Coffee'
    ws.cell(row=6, column=3).value = 'Drift Coffee Co.'
    wb.create_sheet('Additional')
    return wb

_suma_p0 = {
    'product_name': 'Ancient Defence', 'brand_name': 'Ancestral Superfoods',
    'sku_code': 'F-ADEF-100', 'allergen_details': [], 'certifications': [],
}
_suma_p1 = {
    'product_name': 'Original Cold Brew Coffee', 'brand_name': 'Drift Coffee Co.',
    'sku_code': 'DCC-CBR-ORI-250', 'allergen_details': [], 'certifications': [],
}
_suma_buf = io.BytesIO()
_build_suma_tabs_wb().save(_suma_buf)
_suma_precomputed = [
    {'sheet_name': 'Product 1', 'row': 4, 'col': 3, 'value': 'F-ADEF-100',
     'field_label': 'Suma Product Code', 'product_index': 0},
    {'sheet_name': 'Product 1', 'row': 5, 'col': 3, 'value': 'Original Cold Brew Coffee',
     'field_label': 'Full Product Name', 'product_index': 1},
    {'sheet_name': 'Product 1', 'row': 6, 'col': 3, 'value': 'Drift Coffee Co.',
     'field_label': 'Brand', 'product_index': 1},
]
req_suma = FillRequest(
    file_base64=base64.b64encode(_suma_buf.getvalue()).decode(),
    products=[_suma_p0, _suma_p1],
    retailer_name='Suma',
    fill_mode='tabs',
    form_spec={'layout': 'vertical', 'data_sheet': 'Product 1', 'label_column': 2, 'value_column': 3},
    precomputed_mappings=_suma_precomputed,
)
res_suma = asyncio.run(fill_nlf(req_suma))
wb_suma = openpyxl.load_workbook(io.BytesIO(base64.b64decode(res_suma['file_base64'])))
ws_p1 = wb_suma['Product 1']
ws_p2 = wb_suma['Product 2']
tabs_clean = (
    ws_p1.cell(row=4, column=3).value in (None, '')
    and ws_p1.cell(row=7, column=3).value == 'F-ADEF-100'
    and ws_p1.cell(row=5, column=3).value == 'Ancient Defence'
    and ws_p1.cell(row=6, column=3).value == 'Ancestral Superfoods'
    and ws_p2.cell(row=5, column=3).value == 'Original Cold Brew Coffee'
    and ws_p2.cell(row=6, column=3).value == 'Drift Coffee Co.'
)
print(f'{"PASS" if tabs_clean else "FAIL"} | Tabs Product 1: retailer code blank, supplier ref + identity consistent')
results.append(tabs_clean)

packaging_desc_ok = (
    _m.map_field('Packaging Description - Item/Case', {
        'product_description': 'Organic Functional Tea Blend',
        'inner_packaging_material': 'Stand-up pouch',
        'case_size_description': '12 x 100g',
    }) == 'Stand-up pouch'
    and _m.map_field('Packaging Description', {
        'product_description': 'Organic Functional Tea Blend',
    }) == 'N/A'
)
print(f'{"PASS" if packaging_desc_ok else "FAIL"} | Packaging Description never maps to product_description')
results.append(packaging_desc_ok)

retailer_code_ok = (
    _m.map_field('Suma Product Code', TEST_PRODUCT) is None
    and _m.map_field('Ocado Article Number', TEST_PRODUCT) is None
    and _m.map_field('Product Code', TEST_PRODUCT) is None
    and _m.map_field('Supplier reference code', TEST_PRODUCT) == 'TB-001'
    and _m.map_field('Supplier Product Code', TEST_PRODUCT) == 'TB-001'
    and _m.is_retailer_owned_code_field('Suma Product Code')
    and not _m.is_retailer_owned_code_field('Supplier reference code')
)
print(f'{"PASS" if retailer_code_ok else "FAIL"} | Retailer product codes blank; supplier ref → SKU')
results.append(retailer_code_ok)

# 5-product tabs: every identity field on Product 1 tab must belong to products[0]
def _build_tabs_wb_n(n):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Product 1'
    ws.cell(row=4, column=2).value = 'Suma Product Code'
    ws.cell(row=5, column=2).value = 'Full Product Name'
    ws.cell(row=6, column=2).value = 'Brand name'
    ws.cell(row=7, column=2).value = 'Supplier reference code'
    ws.cell(row=54, column=2).value = 'Packaging Description - Item/Case'
    wb.create_sheet('Notes')
    return wb

_tabs5 = [
    {
        'product_name': f'Ancestral Product {i}',
        'brand_name': f'Ancestral Brand {i}',
        'sku_code': f'F-SKU-{i}',
        'product_description': f'Organic Raw Activated description {i}',
        'inner_packaging_material': f'Pouch type {i}',
        'allergen_details': [], 'certifications': [],
    }
    for i in range(5)
]
_buf5 = io.BytesIO()
_build_tabs_wb_n(5).save(_buf5)
# Simulate stale __shared__ contamination: product 2 name/brand in index-0 precomputed
_contaminated_pre = []
for idx, p in enumerate(_tabs5):
    _contaminated_pre.extend([
        {'sheet_name': f'Product {idx + 1}', 'row': 4, 'col': 3,
         'value': p['sku_code'], 'field_label': 'Suma Product Code', 'product_index': idx},
        {'sheet_name': f'Product {idx + 1}', 'row': 5, 'col': 3,
         'value': p['product_name'], 'field_label': 'Full Product Name', 'product_index': idx},
        {'sheet_name': f'Product {idx + 1}', 'row': 6, 'col': 3,
         'value': p['brand_name'], 'field_label': 'Brand name', 'product_index': idx},
    ])
# Old bug: product 1's name leaked onto Product 1 tab precomputed index 0
_contaminated_pre.append({
    'sheet_name': 'Product 1', 'row': 5, 'col': 3,
    'value': 'Pure Pea Protein WRONG', 'field_label': 'Full Product Name', 'product_index': 0,
})
_contaminated_pre.append({
    'sheet_name': 'Product 1', 'row': 6, 'col': 3,
    'value': 'Form Nutrition WRONG', 'field_label': 'Brand name', 'product_index': 0,
})
req5 = FillRequest(
    file_base64=base64.b64encode(_buf5.getvalue()).decode(),
    products=_tabs5,
    retailer_name='Suma',
    fill_mode='tabs',
    form_spec={'layout': 'vertical', 'data_sheet': 'Product 1', 'label_column': 2, 'value_column': 3},
    precomputed_mappings=_contaminated_pre,
)
res5 = asyncio.run(fill_nlf(req5))
wb5 = openpyxl.load_workbook(io.BytesIO(base64.b64decode(res5['file_base64'])))
ws5_p1 = wb5['Product 1']
tabs5_ok = True
for tab_i in range(5):
    ws_t = wb5[f'Product {tab_i + 1}']
    p = _tabs5[tab_i]
    if ws_t.cell(row=4, column=3).value not in (None, ''):
        tabs5_ok = False
    if ws_t.cell(row=7, column=3).value != p['sku_code']:
        tabs5_ok = False
    if ws_t.cell(row=5, column=3).value != p['product_name']:
        tabs5_ok = False
    if ws_t.cell(row=6, column=3).value != p['brand_name']:
        tabs5_ok = False
    pkg = ws_t.cell(row=54, column=3).value
    if pkg == p['product_description']:
        tabs5_ok = False
print(f'{"PASS" if tabs5_ok else "FAIL"} | 5-product tabs: identity consistent + packaging != description')
results.append(tabs5_ok)

print()
print('--- Missing API key / Claude mock tests ---')

import main as main_mod
from unittest.mock import patch

# Known fields fill without Anthropic; unknown stays blank
with patch.object(main_mod, 'get_anthropic_api_key', return_value=None):
    resolved_no_key = main_mod.resolve_values_for_labels(
        ['THR Licensed', 'Brand Name', 'Comparative Unit'], TEST_PRODUCT,
    )
no_key_ok = (
    resolved_no_key.get('Brand Name') == 'TestBrand'
    and 'THR Licensed' not in resolved_no_key
    and 'Comparative Unit' not in resolved_no_key
)
print(f'{"PASS" if no_key_ok else "FAIL"} | Without API key: known filled, unknown blank')
results.append(no_key_ok)

# Mock Claude resolves unknown label when key would be used
with patch.object(main_mod, 'claude_resolve_fields', return_value={'THR Licensed': 'No'}):
    resolved_mock = main_mod.resolve_values_for_labels(['THR Licensed', 'Brand Name'], TEST_PRODUCT)
mock_ok = resolved_mock.get('THR Licensed') == 'No' and resolved_mock.get('Brand Name') == 'TestBrand'
print(f'{"PASS" if mock_ok else "FAIL"} | Mock Claude resolves unknown label')
results.append(mock_ok)

print()
passed = sum(1 for r in results if r is True)
total = len(results)
print(f'RESULT: {passed}/{total} tests passed')
if passed < total:
    print('TESTS FAILED — DO NOT DEPLOY')
    exit(1)
else:
    print('ALL TESTS PASSED — SAFE TO DEPLOY')
