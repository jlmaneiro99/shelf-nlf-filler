from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import List, Any, Dict, Optional
import base64
import io
import json
import openpyxl
from openpyxl.cell.cell import MergedCell
from copy import copy, deepcopy
import datetime
import os
import re

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

import httpx

app = FastAPI()
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


class ProductData(BaseModel):
    data: Dict[str, Any]


class FillRequest(BaseModel):
    file_base64: str
    products: List[Dict[str, Any]]
    retailer_name: str
    fill_mode: str = 'auto'
    form_spec: Optional[Dict[str, Any]] = None
    precomputed_mappings: Optional[List[Dict[str, Any]]] = None


def get_anthropic_api_key() -> Optional[str]:
    """Read at call time so Railway env vars apply after redeploy without stale cache."""
    key = os.environ.get('ANTHROPIC_API_KEY')
    if key and str(key).strip():
        return str(key).strip()
    return None


def anthropic_config_status() -> Dict[str, Any]:
    key = get_anthropic_api_key()
    httpx_ok = False
    try:
        import httpx as _httpx  # noqa: F401
        httpx_ok = True
    except ImportError:
        pass
    # Names only — helps detect typos like CLAUDE_API_KEY without exposing values
    related_env_keys = sorted(
        k for k in os.environ
        if any(x in k.upper() for x in ('ANTHROPIC', 'CLAUDE', 'OPENAI'))
    )
    return {
        'anthropic_key_present': bool(key),
        'anthropic_key_source': 'env' if key else None,
        'anthropic_client_available': httpx_ok,
        'related_env_keys_present': related_env_keys,
    }


class _FillGuard:
    """Tracks blocked formula writes during a single /fill request."""
    formula_blocks = 0

    @classmethod
    def reset(cls):
        cls.formula_blocks = 0


class _WriteTracker:
    """Records every (sheet, row, col) written so a single /fill request can be
    audited for conflicting layout passes (e.g. a stray vertical dump mixed into
    a horizontal fill)."""
    writes = []
    enabled = False

    @classmethod
    def reset(cls):
        cls.writes = []
        cls.enabled = True

    @classmethod
    def record(cls, sheet, row, col):
        if cls.enabled:
            cls.writes.append((sheet, int(row), int(col)))


def is_formula_cell(cell):
    if isinstance(cell, MergedCell):
        return False
    if cell.data_type == 'f':
        return True
    v = cell.value
    return isinstance(v, str) and str(v).startswith('=')


def count_formula_cells(wb):
    count = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell, MergedCell) and is_formula_cell(cell):
                    count += 1
    return count


def safe_write(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for rng in ws.merged_cells.ranges:
            if (rng.min_row <= row <= rng.max_row and
                    rng.min_col <= col <= rng.max_col):
                target = ws.cell(row=rng.min_row, column=rng.min_col)
                if is_formula_cell(target):
                    _FillGuard.formula_blocks += 1
                    return False
                if len(str(value)) > 500:
                    value = str(value)[:500]
                target.value = value
                _WriteTracker.record(ws.title, target.row, target.column)
                return True
        return False
    if is_formula_cell(cell):
        _FillGuard.formula_blocks += 1
        return False
    if len(str(value)) > 500:
        value = str(value)[:500]
    cell.value = value
    _WriteTracker.record(ws.title, row, col)
    return True


ALLERGEN_KEYS = [
    ('gluten', ['gluten', 'cereal', 'wheat', 'rye', 'barley', 'oat']),
    ('eggs', ['egg']),
    ('fish', ['fish']),
    ('crustaceans', ['crustacean', 'shellfish', 'prawn', 'shrimp', 'crab', 'lobster']),
    ('molluscs', ['mollusc', 'mussel', 'oyster', 'clam', 'squid']),
    ('peanuts', ['peanut', 'groundnut']),
    ('soybeans', ['soy', 'soya', 'soybean']),
    ('milk', ['milk', 'dairy', 'lactose', 'whey', 'casein']),
    ('nuts', ['nut', 'almond', 'hazelnut', 'walnut', 'cashew', 'pecan', 'pistachio', 'macadamia', 'brazil']),
    ('celery', ['celery', 'celeriac']),
    ('mustard', ['mustard']),
    ('sesame', ['sesame']),
    ('sulphites', ['sulphite', 'sulphur', 'sulfite', 'sulfur', 'so2']),
    ('lupin', ['lupin', 'lupine']),
    ('royal jelly', ['royal jelly']),
    ('propolis', ['propolis']),
    ('bee pollen', ['bee pollen']),
]


def safe_str(value, default='N/A'):
    if value is None:
        return default
    s = str(value).strip()
    if s == '' or s == 'None' or s == 'null':
        return default
    return s


def norm_label(label):
    """Normalise a field label for matching: lower, collapse whitespace, drop trailing *."""
    if label is None:
        return ''
    return re.sub(r'\s+', ' ', str(label).lower().replace('*', ' ')).strip()


def is_junk_header(label):
    """A header cell that is not a real field: blank, purely numeric (e.g. corrupted
    merged value '1.65'), or shorter than 3 characters."""
    if label is None:
        return True
    s = str(label).strip()
    if len(s) < 3:
        return True
    cleaned = s.replace(',', '').replace('%', '').replace('£', '').replace('$', '').strip()
    if re.fullmatch(r'-?\d+(\.\d+)?', cleaned):
        return True
    return False


def is_formula_computed_field(label):
    """Fields the form computes itself via a formula — never write a value directly."""
    n = norm_label(label)
    return 'margin' in n


BARE_PACKAGING_HEADERS = frozenset({
    'other', 'paper', 'glass', 'plastic', 'card', 'cardboard',
    'metal', 'aluminium', 'aluminum', 'steel', 'tin', 'wood', 'none',
})

RETAILER_CODE_QUALIFIERS = frozenset({
    'suma', 'ocado', 'tesco', 'sainsbury', 'waitrose', 'morrisons', 'asda',
    'dunnes', 'dundeis', 'clf', 'spark', 'boots', 'costco', 'whole foods',
    'wholefoods', 'trader joe', 'kroger', 'walmart', 'heb', 'loblaws', 'coles',
    'woolworths', 'amazon', 'delhaize', 'metro', 'carrefour', 'aldi', 'lidl',
    'target', 'safeway',
})


def is_supplier_code_field(label):
    """Supplier/vendor SKU fields — safe to fill with sku_code."""
    n = norm_label(label)
    if any(x in n for x in (
        'supplier product code', 'supplier reference', 'supplier ref',
        'supplier code', 'supplier sku', 'your code', 'vendor code',
        'vendor ref', 'supplier item code',
    )):
        return 'retailer' not in n and 'etail' not in n and 'buyer' not in n
    if re.search(r'\bsku\b', n) and not any(x in n for x in ('retailer', 'etail', 'buyer', 'suma', 'ocado', 'tesco')):
        return True
    if re.search(r'\bsku\s*/\s*supplier\b', n):
        return True
    return False


def is_retailer_owned_code_field(label):
    """Retailer/buyer-assigned codes — supplier leaves blank (Suma Product Code, Tesco TPN, etc.)."""
    if is_supplier_code_field(label):
        return False
    n = norm_label(label)
    if any(x in n for x in (
        'etail code', 'retailer code', 'retailer sku', 'retailer product',
        'buyer code', 'buyer sku',
    )):
        return True
    if re.search(r'\btpn\b', n) and not any(x in n for x in ('supplier', 'vendor', 'your')):
        return True
    if any(x in n for x in ('product code', 'article number', 'article no')):
        if not any(x in n for x in ('supplier', 'vendor', 'your')):
            return True
    for retailer in RETAILER_CODE_QUALIFIERS:
        if retailer in n and re.search(r'code|article|tpn|sku|number|ref', n):
            if not any(x in n for x in ('supplier', 'vendor', 'your')):
                return True
    return False


def is_bare_packaging_material_header(label):
    """Single-word packaging material column headers — not product data fields."""
    n = norm_label(label)
    if n in BARE_PACKAGING_HEADERS:
        return True
    first = n.split()[0] if n else ''
    if first in BARE_PACKAGING_HEADERS:
        return True
    return bool(re.match(
        r'^(other|paper|glass|plastic|card|cardboard|metal|aluminium|aluminum|steel|tin|wood|none)(\*|\s|\(|$)',
        n,
    ))


def unit_uom(p):
    wu = safe_str(p.get('weight_unit'), '').lower()
    if wu in ('g', 'ml'):
        return wu
    if wu == 'kg':
        return 'g'
    if wu in ('l', 'litre', 'liter', 'liters', 'litres'):
        return 'ml'
    fmt = safe_str(p.get('product_format'), '').lower()
    if fmt in ('liquid', 'drink', 'beverage'):
        return 'ml'
    return 'g'


def trade_price_per_unit(p):
    trade = p.get('trade_price_per_case')
    units = p.get('units_per_case')
    if trade is None or not units:
        return None
    try:
        v = round(float(trade) / float(units), 4)
        if v == int(v):
            return str(int(v))
        return str(v)
    except (TypeError, ValueError, ZeroDivisionError):
        return None


def cost_price_per_unit(p):
    cost = p.get('cost_price_per_case')
    units = p.get('units_per_case')
    if cost is None or not units:
        return None
    try:
        v = round(float(cost) / float(units), 4)
        if v == int(v):
            return str(int(v))
        return str(v)
    except (TypeError, ValueError, ZeroDivisionError):
        return None


def _format_numeric_size(val):
    try:
        fv = float(val)
        if fv == int(fv):
            return str(int(fv))
        return str(val).strip()
    except (TypeError, ValueError):
        return safe_str(val, '')


def parse_unit_size_from_product(p):
    """Numeric unit size from unit_net_weight_g, unit_size alias, or variant (e.g. 40g)."""
    uw = p.get('unit_net_weight_g')
    if uw is not None and str(uw).strip() not in ('', 'None', 'null'):
        formatted = _format_numeric_size(uw)
        if formatted:
            return formatted
    us = p.get('unit_size')
    if us is not None and str(us).strip() not in ('', 'None', 'null'):
        formatted = _format_numeric_size(us)
        if formatted:
            return formatted
    variant = safe_str(p.get('variant'), '').strip()
    if variant:
        m = re.match(r'^(\d+(?:\.\d+)?)\s*([a-zA-Z]+)?\s*$', variant.replace(' ', ''))
        if m:
            return _format_numeric_size(m.group(1))
    return ''


def parse_uom_from_product(p):
    """Unit of measure from weight_unit or variant suffix (40g → g)."""
    variant = safe_str(p.get('variant'), '').strip().replace(' ', '')
    if variant:
        m = re.match(r'^\d+(?:\.\d+)?([a-zA-Z]+)$', variant)
        if m:
            u = m.group(1).lower()
            if u == 'kg':
                return 'g'
            if u in ('l', 'litre', 'liter', 'ltr', 'liters', 'litres'):
                return 'ml'
            if u in ('g', 'ml'):
                return u
    return unit_uom(p)


def is_trade_case_cost_label(label):
    if 'trade case cost' in label:
        return True
    return 'case cost' in label and 'trade' in label


def is_trade_unit_cost_label(label):
    if 'trade unit cost' in label:
        return True
    return 'trade' in label and 'unit cost' in label


def is_supplier_case_cost_label(label):
    if 'dundeis case cost' in label:
        return True
    return 'case cost' in label and any(x in label for x in ('dundeis', 'supplier', 'landed'))


def is_supplier_unit_cost_label(label):
    if 'dundeis unit cost' in label:
        return True
    return 'unit cost' in label and any(x in label for x in ('dundeis', 'supplier', 'landed'))


def has_cert(p, *keywords):
    certs = [str(c).lower() for c in (p.get('certifications') or [])]
    return any(all(k in c for k in keywords) for c in certs)


def claude_resolve_fields(unresolved_labels, product):
    api_key = get_anthropic_api_key()
    if not api_key or not unresolved_labels:
        return {}
    unique_labels = list(dict.fromkeys(unresolved_labels))
    prompt = f'''You are an expert FMCG sales manager filling a retailer New Line Form. For each form field label below, return the correct value from the product data, exactly as an experienced human would write it.

PRODUCT DATA:
{json.dumps(product, default=str)}

FORM FIELDS TO RESOLVE:
{json.dumps(unique_labels)}

Rules:
- Yes/No questions get "Yes" or "No"
- Unknown/not applicable gets "N/A"
- Numbers get plain numbers, no units
- Never invent data that is not in the product data
- If the field asks about licensing, registration, or legal status not in the data, answer "No" or "N/A" conservatively — never claim a certification the product does not have
- Unit conversions: oz = g/28.35, lbs = kg*2.205, sodium mg = salt_g/2.5*1000

Return ONLY JSON: {{"label": "value", ...}}'''

    try:
        resp = httpx.post(
            'https://api.anthropic.com/v1/messages',
            headers={
                'x-api-key': api_key,
                'anthropic-version': '2023-06-01',
                'content-type': 'application/json',
            },
            json={
                'model': 'claude-sonnet-4-20250514',
                'max_tokens': 4000,
                'temperature': 0,
                'messages': [{'role': 'user', 'content': prompt}],
            },
            timeout=60,
        )
        if resp.status_code != 200:
            return {}
        data = resp.json()
        content = data.get('content') or []
        if not content:
            return {}
        text = content[0].get('text', '')
        text = text.strip()
        if text.startswith('```json'):
            text = text[7:]
        elif text.startswith('```'):
            text = text[3:]
        if text.endswith('```'):
            text = text[:-3]
        text = text.strip()
        return json.loads(text)
    except Exception:
        return {}


def resolve_values_for_labels(labels, product):
    """Layer 2 + 3: rules first, one Claude batch for unresolved labels."""
    resolved = {}
    unresolved = []
    for label in labels:
        v = map_field(label, product)
        if v is not None:
            if v in ('None', 'null', ''):
                v = 'N/A'
            resolved[label] = v
        else:
            unresolved.append(label)
    if unresolved:
        unresolved = [l for l in unresolved if not is_bare_packaging_material_header(l)]
    if unresolved:
        claude_vals = claude_resolve_fields(unresolved, product)
        for label in unresolved:
            v = claude_vals.get(label)
            if v is not None:
                if v in ('None', 'null', ''):
                    v = 'N/A'
                resolved[label] = v
    return resolved


def get_allergen_value(label, allergen_details):
    if not allergen_details:
        return 'Not Present'
    label_lower = label.lower()
    for category, keywords in ALLERGEN_KEYS:
        if any(kw in label_lower for kw in keywords):
            for a in allergen_details:
                allergen_name = str(a.get('allergen', '')).lower()
                if any(kw in allergen_name for kw in keywords):
                    if a.get('present_in_formulation') or a.get('present'):
                        return 'Present'
                    if a.get('may_contain_traces') or a.get('may_contain'):
                        return 'May Contain'
            return 'Not Present'
    return 'Not Present'


def is_allergen_field(label):
    label_lower = label.lower()
    if 'gluten free' in label_lower or 'gluten-free' in label_lower:
        return False
    if 'free from' in label_lower:
        return False
    if 'dairy free' in label_lower or 'nut free' in label_lower:
        return False
    if 'soy free' in label_lower or 'soya free' in label_lower:
        return False
    for _, keywords in ALLERGEN_KEYS:
        if any(kw in label_lower for kw in keywords):
            return True
    return False


def nutritional_value(val, label, serving_size):
    if val is None:
        return None
    label_lower = label.lower()
    v = float(val)
    if 'per serving' in label_lower and serving_size:
        v = round(v * float(serving_size) / 100, 1)
    elif 'per 50g' in label_lower:
        v = round(v * 0.5, 1)
    elif 'per 30g' in label_lower:
        v = round(v * 0.3, 1)
    elif 'per 25g' in label_lower:
        v = round(v * 0.25, 1)
    else:
        v = round(v, 2)
        if v == int(v):
            v = int(v)
    return str(v)


def map_field(label_raw, product):
    label = label_raw.lower().strip().rstrip('*').strip()
    p = product
    serving = p.get('serving_size_value')

    if is_allergen_field(label):
        return get_allergen_value(label, p.get('allergen_details', []))

    if 'packaging description' in label:
        v = safe_str(p.get('inner_packaging_material'))
        if not v:
            v = safe_str(p.get('case_size_description'))
        return v if v else 'N/A'

    # Bare packaging headers before any weight/size matchers (never leak inner_plastic etc.).
    if is_bare_packaging_material_header(label):
        material = norm_label(label).split()[0]
        explicit = p.get(f'packaging_{material}_weight_g') or p.get(f'packaging_{material}')
        if explicit:
            return safe_str(explicit)
        return 'N/A'

    if re.search(r'\bproduct name\b', label) or label in (
        'name of product', 'article name', 'item name', 'product title',
        'full product name', 'product name on pack', 'product name *',
    ):
        return safe_str(p.get('product_name'), '')

    if label in ['brand name', 'brand', 'brand / manufacturer',
                 'manufacturer', 'manufacturer name', 'brand name *']:
        return safe_str(p.get('brand_name'), '')

    if any(x in label for x in ['supplier name', 'supplier / company',
                                  'company name', 'vendor name',
                                  'supplier company']):
        return safe_str(p.get('supplier_name') or p.get('brand_name'), '')

    # Customs/tariff codes must NEVER receive the SKU — handle them first.
    if any(x in label for x in ['commodity code', 'commodity', 'hs code', 'hs/commodity',
                                  'hs / commodity', 'hts code', 'tariff code', 'tariff',
                                  'customs code', 'import code', 'meursing']):
        if 'meursing' in label:
            return safe_str(p.get('meursing_code'))
        return safe_str(p.get('hs_commodity_code'))

    if is_retailer_owned_code_field(label):
        return None

    if is_supplier_code_field(label):
        return safe_str(p.get('sku_code'), '')

    if any(x in label for x in ['case barcode', 'outer barcode',
                                  'case ean', 'shipper barcode',
                                  'carton barcode', 'case gtin',
                                  'case upc', 'outer ean',
                                  'upc barcode case', 'barcode case',
                                  'barcode — case']):
        return safe_str(p.get('case_barcode'))

    if any(x in label for x in ['ean', 'barcode', 'gtin', 'upc',
                                  'unit barcode', 'individual barcode',
                                  'individual unit barcode',
                                  'product barcode', 'item barcode',
                                  'unit ean', 'unit gtin']):
        return safe_str(p.get('ean_barcode'), '')

    if any(x in label for x in ['variant', 'pack size', 'product size',
                                  'size / format', 'format', 'pack format']):
        return safe_str(p.get('variant'), '')

    if ('product description' in label or re.search(r'\bdescription\b', label)) and \
       'product name' not in label and 'usp' not in label and 'sell' not in label and \
       'packaging description' not in label:
        return safe_str(p.get('product_description'), '')

    if any(x in label for x in ['usp', 'key claims', 'unique selling',
                                  'selling point', 'key benefit',
                                  'product claim', 'about the product',
                                  'sell copy', 'marketing description',
                                  'consumer description', 'website description']):
        return safe_str(p.get('usp'), '')

    if 'ingredient' in label:
        return safe_str(p.get('ingredients'))

    # Unit Size / UOM — Dundeis, BP, and similar horizontal forms
    if label in ('unit size', 'unit size *') or (
        'unit size' in label and 'case' not in label and 'uom' not in label
    ):
        size = parse_unit_size_from_product(p)
        return size if size else 'N/A'

    if label in ('uom', 'uom *', 'unit uom', 'unit of measure', 'unit of measurement') or \
       label.startswith('uom '):
        return parse_uom_from_product(p)

    # Trade / cost columns — label-specific (before generic wholesale/cost matchers)
    if is_trade_case_cost_label(label):
        v = p.get('trade_price_per_case')
        return safe_str(v, 'N/A') if v is not None else 'N/A'

    if is_trade_unit_cost_label(label):
        v = trade_price_per_unit(p)
        return safe_str(v, 'N/A') if v is not None else 'N/A'

    if is_supplier_case_cost_label(label):
        v = p.get('cost_price_per_case')
        return safe_str(v, 'N/A') if v is not None else 'N/A'

    if is_supplier_unit_cost_label(label):
        v = cost_price_per_unit(p)
        return safe_str(v, 'N/A') if v is not None else 'N/A'

    if any(x in label for x in ['how will you promote', 'promotional plan',
                                  'promotional support', 'promotional activity',
                                  'marketing support', 'trade support',
                                  'marketing plan', 'consumer marketing',
                                  'promotion plan', 'how do you plan to market']):
        return safe_str(p.get('promotion_plan'))

    if any(x in label for x in ['rrp', 'retail price', 'recommended retail',
                                  'msrp', 'normal rrp', 'consumer price',
                                  'selling price', 'shelf price']):
        return safe_str(p.get('rrp'), '')

    if any(x in label for x in ['wholesale price', 'trade price',
                                  'cost to', 'normal trade price',
                                  'invoice price', 'supply price',
                                  'ex-works price', 'unit cost to retailer',
                                  'buying price', 'net price']):
        return safe_str(p.get('trade_price_per_case'), '')

    if any(x in label for x in ['cost price', 'landed cost',
                                  'our cost']) and 'wholesale' not in label:
        return safe_str(p.get('cost_price_per_case'))

    if any(x in label for x in ['units per case', 'units/case',
                                  'case quantity', 'qty per case',
                                  'pieces per case', 'count per case',
                                  'units per outer']):
        return safe_str(p.get('units_per_case'), '')

    if any(x in label for x in ['case size description', 'case configuration',
                                  'pack configuration', 'pack description',
                                  'case contents', 'case format']) or \
       ('case size' in label and any(x in label for x in ['eg', 'e.g', 'example', 'x'])):
        return safe_str(p.get('case_size_description'), '')

    if any(x in label for x in ['vat rate', 'vat', 'tax rate', 'gst rate',
                                  'hst', 'sales tax', 'tax / vat',
                                  'tax type']):
        return safe_str(p.get('vat_rate'), '')

    if any(x in label for x in ['minimum order', 'moq', 'min order',
                                  'minimum quantity']):
        return safe_str(p.get('moq_units') or p.get('moq_value'))

    if any(x in label for x in ['lead time', 'delivery time',
                                  'lead time (days)', 'lead time (weeks)',
                                  'turnaround time']):
        return safe_str(p.get('lead_time_days'))

    if any(x in label for x in ['payment terms', 'terms of payment',
                                  'credit terms', 'payment conditions']):
        return safe_str(p.get('payment_terms'))

    if any(x in label for x in ['case gross weight', 'gross weight',
                                  'gross case weight', 'total case weight',
                                  'case weight (gross)', 'weight incl']):
        return safe_str(p.get('case_gross_weight_kg'))

    if any(x in label for x in ['case net weight', 'net weight',
                                  'net case weight', 'product weight only',
                                  'case weight (net)', 'weight excl']):
        return safe_str(p.get('case_net_weight_kg'))

    if any(x in label for x in ['unit weight', 'individual unit weight',
                                  'net weight per unit', 'unit net weight',
                                  'product weight per unit']):
        return safe_str(p.get('unit_net_weight_g'))

    if any(x in label for x in ['minimum shelf life', 'min shelf life',
                                  'shelf life on delivery', 'shelf life at receipt',
                                  'minimum remaining shelf life',
                                  'guaranteed shelf life']):
        v = p.get('min_shelf_life_on_delivery_weeks')
        if v:
            if 'day' in label:
                return str(int(v) * 7)
            return str(v)
        return 'N/A'

    if any(x in label for x in ['shelf life', 'best before', 'product life',
                                  'life of product', 'expiry', 'use by',
                                  'total shelf life']):
        v = p.get('shelf_life_weeks')
        if v:
            if 'day' in label:
                return str(int(v) * 7)
            return str(v)
        return 'N/A'

    if any(x in label for x in ['storage conditions', 'storage criteria',
                                  'storage requirement', 'store']):
        return safe_str(p.get('storage_conditions'), '')

    if any(x in label for x in ['storage instructions', 'storage advice',
                                  'how to store', 'storage guidance']):
        return safe_str(p.get('storage_instructions'))

    if any(x in label for x in ['cases per pallet', 'pallet configuration',
                                  'cases/pallet', 'units per pallet']):
        return safe_str(p.get('cases_per_pallet'))

    if 'cases per layer' in label or 'layers per pallet' in label:
        return safe_str(p.get('cases_per_layer') or p.get('layers_per_pallet'))

    if any(x in label for x in ['country of provenance', 'provenance',
                                  'last country of duty', 'duty paid country']):
        return safe_str(p.get('country_of_provenance') or p.get('country_of_origin'), '')

    if any(x in label for x in ['country of origin', 'country of manufacture',
                                  'country of production', 'made in',
                                  'produced in', 'origin country',
                                  'place of manufacture']):
        return safe_str(p.get('country_of_origin'), '')

    if any(x in label for x in ['hs code', 'hs / commodity', 'tariff code',
                                  'commodity code', 'customs code',
                                  'hts code', 'import code']) and 'sku' not in label:
        return safe_str(p.get('hs_commodity_code'))

    if 'meursing' in label:
        return safe_str(p.get('meursing_code'))

    if any(x in label for x in ['eu address', 'european address']):
        return 'Yes' if p.get('eu_address_on_pack') else 'No'

    if any(x in label for x in ['uk address', 'united kingdom address',
                                  'gb address']):
        return 'Yes' if p.get('uk_address_on_pack') else 'No'

    if 'vegan' in label and 'non' not in label:
        return 'Yes' if p.get('is_vegan') else 'No'

    if 'vegetarian' in label and 'non' not in label:
        return 'Yes' if p.get('is_vegetarian') else 'No'

    if any(x in label for x in ['gluten free', 'gluten-free', 'gf ',
                                  ' gf', 'free from gluten']):
        return 'Yes' if p.get('is_gluten_free') else 'No'

    if any(x in label for x in ['organic certification number',
                                  'organic cert number', 'cert number',
                                  'certification number', 'accreditation number',
                                  'approval number', 'organic body']):
        return safe_str(p.get('organic_cert_number'))

    if 'nasaa' in label:
        return 'Yes' if has_cert(p, 'nasaa') else 'No'

    if any(x in label for x in ['australian certified organic']) or label.strip() == 'aco':
        return 'Yes' if has_cert(p, 'aco') or has_cert(p, 'australian', 'organic') else 'No'

    if 'soil association' in label:
        return 'Yes' if has_cert(p, 'soil') else 'No'

    if any(x in label for x in ['usda organic', 'usda certified']):
        return 'Yes' if has_cert(p, 'usda') else 'No'

    if 'organic' in label and 'cert' not in label and 'number' not in label:
        return 'Yes' if p.get('is_organic') else 'No'

    if any(x in label for x in ['fairtrade', 'fair trade', 'fair-trade']):
        return 'Yes' if p.get('is_fairtrade') else 'No'

    if any(x in label for x in ['added sugar', 'contains sugar']):
        if not any(u in label for u in ['g per', 'mg per', 'grams', '(g)', 'g)', 'per 100', 'per serving']):
            v = p.get('contains_added_sugar')
            if v is None:
                return 'N/A'
            return 'Yes' if v else 'No'

    if any(x in label for x in ['gm free', 'non-gmo', 'gmo free',
                                  'non gmo', 'genetically modified free',
                                  'not genetically modified']):
        return 'Yes' if p.get('is_gm_free') else 'No'

    if any(x in label for x in ['hfss scope', 'is product hfss',
                                  'hfss in scope']):
        return 'Yes' if p.get('hfss_scope') else 'No'

    if any(x in label for x in ['hfss score', 'nutrient profile score',
                                  'npm score', 'hfss nutrient']):
        return safe_str(p.get('hfss_score'))

    if any(x in label for x in ['less healthy', 'hfss less healthy',
                                  'is product less healthy']):
        return 'Yes' if p.get('hfss_less_healthy') else 'No'

    if 'biodynamic' in label:
        return 'Yes' if p.get('is_biodynamic') else 'No'

    if 'irradiated' in label:
        return 'Yes' if p.get('is_irradiated') else 'No'

    if any(x in label for x in ['contains alcohol', 'alcohol?',
                                  'is this alcoholic', 'alcoholic product']):
        return 'Yes' if p.get('contains_alcohol') else 'No'

    if any(x in label for x in ['abv', 'alcohol by volume',
                                  'alcohol %', 'alcohol content',
                                  'alcohol percentage']):
        return safe_str(p.get('abv_percentage'))

    if 'palm oil free' in label:
        status = str(p.get('palm_oil_status', '')).lower()
        return 'Yes' if 'not contain' in status or 'free' in status else 'No'

    if any(x in label for x in ['palm oil type', 'type of palm oil']):
        return safe_str(p.get('palm_oil_type'))

    if any(x in label for x in ['palm oil percentage', '% palm oil']):
        return safe_str(p.get('palm_oil_percentage'))

    if any(x in label for x in ['palm oil', 'rspo']):
        return safe_str(p.get('palm_oil_status'))

    if any(x in label for x in ['egg status', 'egg sourcing',
                                  'free range egg', 'are they free range']):
        return safe_str(p.get('egg_status'))

    if any(x in label for x in ['dairy free', 'dairy-free',
                                  'free from dairy', 'lactose free']):
        return 'Yes' if p.get('is_dairy_free') else 'No'

    if any(x in label for x in ['soy free', 'soya free',
                                  'free from soy', 'free from soya']):
        v = p.get('allergen_details', [])
        soy_present = any('soy' in str(a.get('allergen', '')).lower()
                         and (a.get('present') or a.get('present_in_formulation'))
                         for a in v)
        return 'No' if soy_present else 'Yes'

    if any(x in label for x in ['nut free', 'free from nuts',
                                  'tree nut free']):
        v = p.get('allergen_details', [])
        nut_present = any('nut' in str(a.get('allergen', '')).lower()
                         and (a.get('present') or a.get('present_in_formulation'))
                         for a in v)
        return 'No' if nut_present else 'Yes'

    if any(x in label for x in ['plant based', 'plant-based',
                                  'suitable for plant based']):
        return 'Yes' if p.get('is_vegan') else 'No'

    if 'halal' in label:
        certs = [str(c).lower() for c in p.get('certifications', [])]
        return 'Yes' if any('halal' in c for c in certs) else 'No'

    if 'kosher' in label:
        certs = [str(c).lower() for c in p.get('certifications', [])]
        return 'Yes' if any('kosher' in c for c in certs) else 'No'

    if any(x in label for x in ['b corp', 'bcorp', 'b-corp']):
        certs = [str(c).lower() for c in p.get('certifications', [])]
        return 'Yes' if any('b corp' in c for c in certs) else 'No'

    if any(x in label for x in ['rainforest alliance',
                                  'rainforest']):
        return 'Yes' if has_cert(p, 'rainforest') else 'No'

    if any(x in label for x in ['non-gmo project', 'non gmo project verified']):
        return 'Yes' if has_cert(p, 'non-gmo') or has_cert(p, 'non gmo') else 'No'

    if 'energy' in label and any(x in label for x in ['kcal', 'cal', 'calorie']):
        return nutritional_value(p.get('energy_kcal'), label, serving)

    if 'energy' in label and any(x in label for x in ['kj', 'kilojoule', 'kilo joule']):
        return nutritional_value(p.get('energy_kj'), label, serving)

    if any(x in label for x in ['energy value', 'energy content',
                                  'calorific value', 'energy/']):
        kcal = p.get('energy_kcal')
        kj = p.get('energy_kj')
        if kcal and kj:
            return f"{kj}kJ / {kcal}kcal"
        return nutritional_value(kcal, label, serving)

    if any(x in label for x in ['saturate', 'saturated fat',
                                  'of which saturate', 'sat fat']):
        return nutritional_value(p.get('saturates'), label, serving)

    if 'monounsaturate' in label:
        return nutritional_value(p.get('monounsaturates'), label, serving)

    if 'polyunsaturate' in label:
        return nutritional_value(p.get('polyunsaturates'), label, serving)

    if 'trans fat' in label or 'trans fatty' in label:
        return '0'

    if any(x in label for x in ['total fat', 'fat content',
                                  'fat g', 'fat per', '- fat',
                                  'of which fat']) or \
       label in ['fat', 'fat *', 'fat (g)', 'fat g per 100g *',
                 'fat g per 100g', 'fat (g per 100g)']:
        return nutritional_value(p.get('fat'), label, serving)

    if any(x in label for x in ['total carbohydrate', 'carbohydrate',
                                  'carbs', 'carbohydrates',
                                  'total carbs', 'of which carbs']):
        return nutritional_value(p.get('carbohydrates'), label, serving)

    if 'added sugar' in label and any(u in label for u in
            ['g per', 'mg per', 'grams', '(g)', 'per 100', 'per serving']):
        if p.get('contains_added_sugar'):
            return nutritional_value(p.get('sugars', 0), label, serving)
        return '0'

    if any(x in label for x in ['total sugar', 'of which sugar',
                                  'sugars', 'sugar content']):
        return nutritional_value(p.get('sugars'), label, serving)

    if any(x in label for x in ['dietary fibre', 'dietary fiber',
                                  'fibre', 'fiber', 'roughage']):
        return nutritional_value(p.get('fibre'), label, serving)

    if any(x in label for x in ['polyol', 'sugar alcohol']):
        return nutritional_value(p.get('polyols'), label, serving)

    if any(x in label for x in ['starch']):
        return nutritional_value(p.get('starch'), label, serving)

    if any(x in label for x in ['total protein', 'protein content',
                                  'protein per', 'of which protein']) or \
       label in ['protein', 'protein *', 'protein g per 100g *',
                 'protein (g per 100g)', 'protein g per 100g']:
        return nutritional_value(p.get('protein'), label, serving)

    if 'sodium' in label:
        salt = p.get('salt')
        if salt is not None:
            serving_size = p.get('serving_size_value')
            sodium_per_100g_mg = float(salt) / 2.5 * 1000
            if 'per serving' in label and serving_size:
                v = round(sodium_per_100g_mg * float(serving_size) / 100, 1)
            elif 'per 100' in label:
                v = round(sodium_per_100g_mg, 1)
            else:
                v = round(sodium_per_100g_mg, 1)
            if v == int(v):
                v = int(v)
            return str(v)
        return 'N/A'

    if any(x in label for x in ['salt equivalent', 'total salt',
                                  'salt content', 'salt per']) or \
       label in ['salt', 'salt *', 'salt g per 100g *',
                 'salt (g per 100g)', 'salt g per 100g']:
        return nutritional_value(p.get('salt'), label, serving)

    if any(x in label for x in ['inner packaging material',
                                  'packaging material', 'packaging type',
                                  'primary packaging']):
        return safe_str(p.get('inner_packaging_material'))

    if any(x in label for x in ['is packaging recyclable', 'recyclable',
                                  'can it be recycled']):
        v = p.get('is_recyclable')
        if v is None:
            return 'N/A'
        return 'Yes' if v else 'No'

    if any(x in label for x in ['biodegradable']):
        return 'Yes' if p.get('is_biodegradable') else 'No'

    if any(x in label for x in ['compostable']):
        return 'Yes' if p.get('is_compostable') else 'No'

    if 'outer packaging' in label and 'plastic' in label:
        return 'Yes' if p.get('outer_packaging_has_plastic') else 'No'

    if 'inner packaging' in label and 'plastic' in label:
        return 'Yes' if p.get('inner_packaging_has_plastic') else 'No'

    if any(x in label for x in ['recycled plastic', '30% recycled',
                                  'recycled content']):
        return 'Yes' if (p.get('inner_plastic_recycled_30pct') or
                         p.get('outer_plastic_recycled_30pct')) else 'No'

    if any(x in label for x in ['fsc', 'pefc', 'paper certified',
                                  'card certified', 'certified paper']):
        return safe_str(p.get('paper_card_certified'), 'Uncertified')

    return None


def detect_layout(ws):
    label_col = 2
    value_col = 3
    header_row = None
    product_col_start = None

    for row in ws.iter_rows(min_row=1, max_row=10):
        for cell in row:
            if cell.value and not isinstance(cell, MergedCell):
                val = str(cell.value).upper()
                if 'PRODUCT 1' in val or 'PRODUCT1' in val:
                    header_row = cell.row
                    product_col_start = cell.column
                    label_col = cell.column - 1
                    value_col = cell.column
                    break
        if header_row:
            break

    return {
        'label_col': label_col,
        'value_col': value_col,
        'header_row': header_row,
        'product_col_start': product_col_start,
        'is_column_format': header_row is not None,
    }


def resolve_sheet_name(wb, sheet_name):
    if not sheet_name:
        return None
    if sheet_name in wb.sheetnames:
        return sheet_name
    target = str(sheet_name).strip().lower()
    for name in wb.sheetnames:
        if name.strip().lower() == target:
            return name
    return None


def find_data_sheet(wb):
    for name in wb.sheetnames:
        nl = name.lower().strip()
        if any(k in nl for k in (
            'product', 'new line', 'nlf', 'submission', 'details', 'form',
        )):
            return name
    best_name = wb.sheetnames[0]
    best_count = 0
    for name in wb.sheetnames:
        det = detect_header_and_layout(wb[name])
        count = det.get('label_count', 0)
        if count > best_count:
            best_count = count
            best_name = name
    return best_name


def detect_header_and_layout(ws):
    best_row = None
    best_count = 0
    for r in range(1, 21):
        count = 0
        for cell in ws[r]:
            if isinstance(cell, MergedCell):
                continue
            v = cell.value
            if isinstance(v, str) and len(v.strip()) > 2 and not v.strip().startswith('='):
                count += 1
        if count > best_count:
            best_count = count
            best_row = r

    if best_row and best_count >= 5:
        first_data = best_row + 1
        row_text = ' '.join(
            str(c.value).lower() for c in ws[first_data]
            if c.value and not isinstance(c, MergedCell)
        )
        if 'example' in row_text or 'sample' in row_text:
            first_data = best_row + 2
        return {
            'layout': 'horizontal_rows',
            'header_row': best_row,
            'first_data_row': first_data,
            'label_count': best_count,
        }
    return {
        'layout': 'vertical',
        'header_row': None,
        'first_data_row': None,
        'label_count': 0,
    }


def count_header_labels(ws, header_row):
    count = 0
    for cell in ws[header_row]:
        if isinstance(cell, MergedCell):
            continue
        if cell.value and not str(cell.value).strip().startswith('='):
            label = str(cell.value).strip()
            if len(label) > 1:
                count += 1
    return count


def col_to_index(col):
    if col is None:
        return None
    if isinstance(col, int):
        return col if col > 0 else None
    s = str(col).strip()
    if not s:
        return None
    if s.isdigit():
        return int(s)
    n = 0
    for ch in s.upper():
        if not ('A' <= ch <= 'Z'):
            return None
        n = n * 26 + (ord(ch) - 64)
    return n if n > 0 else None


def clear_vertical_value_column(ws, value_col, example_rows=None, min_row=1):
    """Remove stale template/example values before filling a vertical tab."""
    example_rows = set(example_rows or [])
    for row in ws.iter_rows(min_row=min_row):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if cell.column != value_col:
                continue
            if cell.row in example_rows:
                continue
            if is_formula_cell(cell):
                continue
            cell.value = None


def sheet_for_precomputed_entry(wb, entry, plan, fill_mode, num_products):
    """Route Step 3 coordinates to the correct tab in tabs-mode vertical fills."""
    product_index = entry.get('product_index')
    try:
        idx = int(product_index) if product_index is not None else 0
    except (TypeError, ValueError):
        idx = 0

    layout = plan.get('layout_used')
    use_tabs = fill_mode in ('tabs',) or (fill_mode == 'auto' and num_products > 1)
    if use_tabs and layout == 'vertical':
        tab_name = f'Product {idx + 1}'
        if tab_name in wb.sheetnames:
            return tab_name

    return resolve_sheet_name(wb, entry.get('sheet_name')) or plan.get('sheet_used')


def apply_precomputed_mappings(wb, mappings, plan, fill_mode='auto', num_products=1):
    """Write Step 3 precomputed field/value pairs using their carried coordinates.
    Only used for NON horizontal_rows layouts (vertical / columns) where the
    coordinates are valid for that layout. Horizontal_rows rebuilds coordinates
    inside fill_horizontal_rows instead."""
    if not mappings:
        return 0
    filled = 0
    for entry in mappings:
        sheet_name = sheet_for_precomputed_entry(wb, entry, plan, fill_mode, num_products)
        if not sheet_name or sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        row = entry.get('row')
        col = col_to_index(entry.get('col'))
        value = entry.get('value')
        if not row or not col or value is None or str(value).strip() == '':
            continue
        if is_formula_computed_field(entry.get('field_label') or entry.get('label')):
            continue  # FIX 5 — never write a margin field directly
        existing = ws.cell(row=int(row), column=int(col))
        if is_formula_cell(existing):  # FIX 5 — formula protection
            continue
        if safe_write(ws, int(row), int(col), str(value)):
            filled += 1
    return filled


def check_write_conflict(plan, num_products):
    """FIX 4 — abort if a single request produced conflicting layout passes
    (e.g. a vertical column dump mixed into a horizontal fill)."""
    from collections import defaultdict
    sheet = plan.get('sheet_used')
    layout = plan.get('layout_used')
    writes = [(r, c) for (s, r, c) in _WriteTracker.writes if s == sheet]
    if not writes:
        return None

    rows_by_col = defaultdict(set)
    cols_by_row = defaultdict(set)
    for r, c in writes:
        rows_by_col[c].add(r)
        cols_by_row[r].add(c)

    if layout == 'horizontal_rows':
        first = plan.get('first_data_row_used')
        if first:
            allowed = {first + i for i in range(max(num_products, 1))}
            stray = sorted({r for (r, c) in writes if r not in allowed})
            wide = any(len(cols) >= 5 for cols in cols_by_row.values())
            tall = any(len(rows) >= 5 for rows in rows_by_col.values())
            if stray and (wide or tall):
                return (
                    f"Layout conflict detected: a horizontal_rows fill also wrote to "
                    f"non-product rows {stray[:10]} on sheet '{sheet}'. A vertical dump "
                    f"was mixed into a horizontal fill — file not returned."
                )
    else:
        vertical = any(len(rows) >= 5 for rows in rows_by_col.values())
        horizontal = any(len(cols) >= 5 for cols in cols_by_row.values())
        if vertical and horizontal:
            return (
                f"Layout conflict detected: both a vertical (one column, many rows) and "
                f"a horizontal (one row, many columns) write pattern occurred on sheet "
                f"'{sheet}' — file not returned."
            )
    return None


def resolve_fill_plan(wb, form_spec, fill_mode):
    spec = form_spec or {}
    sheet_used = resolve_sheet_name(wb, spec.get('data_sheet'))
    layout_used = spec.get('layout')
    header_row_used = spec.get('header_row')
    first_data_row_used = spec.get('first_data_row')
    example_rows = list(spec.get('example_rows') or [])
    other_sheets = set(spec.get('other_sheets') or [])

    if not sheet_used:
        sheet_used = find_data_sheet(wb)

    ws = wb[sheet_used]
    detected = detect_header_and_layout(ws)
    column_layout = detect_layout(ws)

    if fill_mode == 'rows' and not header_row_used:
        best_row, best_count = None, 0
        for r in range(1, 21):
            count = count_header_labels(ws, r)
            if count > best_count:
                best_count, best_row = count, r
        if best_row and best_count >= 3:
            layout_used = 'horizontal_rows'
            header_row_used = best_row
            first_data = best_row + 1
            row_text = ' '.join(
                str(c.value).lower() for c in ws[first_data]
                if c.value and not isinstance(c, MergedCell)
            )
            if 'example' in row_text or 'sample' in row_text:
                first_data = best_row + 2
            first_data_row_used = first_data

    if column_layout['is_column_format'] and layout_used != 'horizontal_rows':
        layout_used = 'horizontal_columns'
    elif detected['layout'] == 'horizontal_rows' and detected.get('label_count', 0) >= 5:
        if not layout_used or layout_used == 'vertical':
            layout_used = 'horizontal_rows'
        if not header_row_used:
            header_row_used = detected['header_row']
        if not first_data_row_used:
            first_data_row_used = detected['first_data_row']
    elif not layout_used:
        layout_used = detected['layout'] or 'vertical'

    if not header_row_used and detected.get('header_row'):
        header_row_used = detected['header_row']
    if not first_data_row_used and detected.get('first_data_row'):
        first_data_row_used = detected['first_data_row']

    labels_found_count = 0
    if layout_used == 'horizontal_rows' and header_row_used:
        labels_found_count = count_header_labels(ws, header_row_used)

    return {
        'sheet_used': sheet_used,
        'layout_used': layout_used,
        'header_row_used': header_row_used,
        'first_data_row_used': first_data_row_used,
        'example_rows': example_rows,
        'other_sheets': other_sheets,
        'labels_found_count': labels_found_count,
        'column_layout': column_layout,
        'label_col': spec.get('label_column') or column_layout['label_col'],
        'value_col': spec.get('value_column') or column_layout['value_col'],
        'first_data_column': spec.get('first_data_column'),
    }


def verify_fill(ws, products, value_col=3, label_col=2):
    """Post-fill sanity checks. Returns list of issues fixed."""
    issues_fixed = []

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if cell.column != value_col:
                continue
            if cell.value is None:
                continue

            val = str(cell.value).strip()
            label_cell = ws.cell(row=cell.row, column=label_col)
            label = str(label_cell.value or '').lower()

            if val in ('None', 'null', 'nan'):
                cell.value = 'N/A'
                issues_fixed.append(f'Row {cell.row}: None → N/A')

            numeric_indicators = ['per 100', 'per serving', 'kcal', 'kj', '(g)', '(kg)', '(mg)',
                                  'weight', 'price', 'rrp', 'cost']
            countries = ['united kingdom', 'ireland', 'united states', 'belgium',
                         'netherlands', 'germany', 'france', 'bulgaria', 'australia']
            if any(ind in label for ind in numeric_indicators):
                if val.lower() in countries:
                    cell.value = None
                    issues_fixed.append(f'Row {cell.row}: country in numeric field removed')

            if any(ind in label for ind in ['g per', 'mg per', 'kcal per', 'kj per']) and 'added' not in label:
                if val in ('Yes', 'No'):
                    cell.value = 'N/A'
                    issues_fixed.append(f'Row {cell.row}: Yes/No in numeric field → N/A')

            barcode_price_indicators = ['barcode', 'ean', 'gtin', 'upc', 'rrp', 'price', 'cost', 'weight']
            if any(ind in label for ind in barcode_price_indicators):
                if val in ('Yes', 'No'):
                    cell.value = 'N/A'
                    issues_fixed.append(f'Row {cell.row}: Yes/No in barcode/price field → N/A')

            if len(val) > 500:
                cell.value = val[:500]
                issues_fixed.append(f'Row {cell.row}: truncated value >500 chars')

    return issues_fixed


def precomputed_identity_conflict(label, pre_val, product):
    """Reject Step 3 precomputed values that clearly belong to a different product."""
    n = norm_label(label)
    val = str(pre_val).strip()
    if not val:
        return False
    pname = safe_str(product.get('product_name'), '')
    brand = safe_str(product.get('brand_name'), '')
    if re.search(r'\bproduct name\b', n) or 'full product name' in n:
        if pname and val != pname and not pname.startswith(val) and val not in pname:
            return True
    if 'brand' in n and brand and val.lower() != brand.lower():
        return True
    return False


def fill_single_sheet(ws, product, value_col=3, label_col=2, example_rows=None,
                      precomputed_by_label=None, product_index=None, debug_tab_fill=False):
    sku = safe_str(product.get('sku_code'), '?')
    product_name = safe_str(product.get('product_name'), '?')
    if debug_tab_fill:
        print(
            f"[TABS_FILL] ENTER sheet={ws.title!r} product_index={product_index} "
            f"sku={sku!r} name={product_name!r} precomputed_fields={len(precomputed_by_label or {})}",
            flush=True,
        )
    print(
        f"[FILL_FN] fill_single_sheet (VERTICAL) ENTER sheet={ws.title!r} "
        f"value_col={value_col} label_col={label_col}",
        flush=True,
    )
    example_rows = set(example_rows or [])
    pre = precomputed_by_label or {}
    fields = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if cell.column != label_col:
                continue
            if cell.row in example_rows:
                continue
            if not cell.value:
                continue
            field_label = str(cell.value).strip()
            if len(field_label) < 3:
                continue
            fields.append((cell.row, field_label))

    labels = [f[1] for f in fields]
    resolved = resolve_values_for_labels(labels, product)

    filled = 0
    for row, field_label in fields:
        if is_retailer_owned_code_field(field_label):
            if debug_tab_fill:
                print(
                    f"[TABS_FILL] SKIP retailer-owned code field label={field_label!r} sku={sku!r}",
                    flush=True,
                )
            continue
        pre_val = pre.get(norm_label(field_label))
        if pre_val is not None and str(pre_val).strip() != '':
            if precomputed_identity_conflict(field_label, pre_val, product):
                if debug_tab_fill:
                    print(
                        f"[TABS_FILL] REJECT precomputed identity conflict "
                        f"label={field_label!r} pre={pre_val!r} sku={sku!r}",
                        flush=True,
                    )
                value = resolved.get(field_label)
                source = 'map_field_guard'
            else:
                value = pre_val
                source = 'precomputed'
        else:
            value = resolved.get(field_label)
            source = 'map_field'
        if value is not None:
            if value in ('None', 'null', ''):
                value = 'N/A'
            if safe_write(ws, row, value_col, value):
                filled += 1
                if debug_tab_fill:
                    print(
                        f"[TABS_FILL] WRITE sheet={ws.title!r} row={row} "
                        f"label={field_label!r} value={value!r} source={source} sku={sku!r}",
                        flush=True,
                    )
    issues_fixed = verify_fill(ws, [product], value_col, label_col)
    return filled, issues_fixed


def build_precomputed_by_label(precomputed):
    """Step 3 mappings → {product_index: {norm_label: value}}.

    For horizontal_rows any row/col coordinates carried by the mapping are
    DISCARDED — coordinates are rebuilt for the horizontal layout. Only the
    field label and value are trusted."""
    by_index = {}
    for entry in precomputed or []:
        label = entry.get('field_label') or entry.get('label')
        value = entry.get('value')
        if not label or value is None or str(value).strip() == '':
            continue
        idx = entry.get('product_index')
        try:
            idx = int(idx) if idx is not None else 0
        except (TypeError, ValueError):
            idx = 0
        by_index.setdefault(idx, {})[norm_label(label)] = str(value)
    return by_index


def fill_horizontal_rows(ws, products, plan, precomputed=None):
    print(f"[FILL_FN] fill_horizontal_rows ENTER sheet={ws.title!r}", flush=True)
    header_row = plan.get('header_row_used')
    first_data_row = plan.get('first_data_row_used')
    if not header_row or not first_data_row:
        return 0, [], plan.get('labels_found_count', 0)

    example_rows = set(plan.get('example_rows') or [])
    pre_by_index = build_precomputed_by_label(precomputed)

    headers = {}
    for cell in ws[header_row]:
        if isinstance(cell, MergedCell):
            continue
        if not cell.value:
            continue
        label = str(cell.value).strip()
        if label.startswith('='):
            continue
        if is_junk_header(label):              # FIX 6 — skip numeric/short junk headers
            continue
        if is_formula_computed_field(label):   # FIX 5 — never write margin etc.
            continue
        headers[cell.column] = label

    labels_found_count = len(headers)
    total_filled = 0
    all_issues = []

    for i, product in enumerate(products):
        target_row = first_data_row + i
        if target_row in example_rows:
            continue
        pre = pre_by_index.get(i, {})
        unresolved = []
        for col, label in headers.items():
            existing = ws.cell(row=target_row, column=col)
            if is_formula_cell(existing):       # FIX 5 — formula protection
                continue
            # Step 3 precomputed value is the source of truth when present and non-empty.
            pre_val = pre.get(norm_label(label))
            if pre_val is not None and str(pre_val).strip() != '':
                value = pre_val
            else:
                value = map_field(label, product)
            if value is not None:
                if value in ('None', 'null', ''):
                    value = 'N/A'
                if safe_write(ws, target_row, col, value):
                    total_filled += 1
            else:
                if not is_bare_packaging_material_header(label):
                    unresolved.append((col, label))
        if unresolved:
            resolved = claude_resolve_fields([l for _, l in unresolved], product)
            for col, label in unresolved:
                if label in resolved and resolved[label]:
                    existing = ws.cell(row=target_row, column=col)
                    if not is_formula_cell(existing):
                        val = resolved[label]
                        if val in ('None', 'null', ''):
                            val = 'N/A'
                        if safe_write(ws, target_row, col, val):
                            total_filled += 1

    return total_filled, all_issues, labels_found_count


def execute_fill(wb, products, plan, fill_mode, precomputed=None):
    ws = wb[plan['sheet_used']]
    column_layout = plan.get('column_layout') or {}
    total_filled = 0
    all_issues = []
    labels_found_count = plan.get('labels_found_count', 0)
    pre_by_index = build_precomputed_by_label(precomputed)

    # ── Resolve the layout ONCE. Never silently fall through to a vertical
    # column dump: if the layout is unknown but a horizontal header row was
    # detected, treat it as horizontal_rows. The vertical writer (fill_single_sheet
    # down a value column) runs ONLY when resolved_layout == 'vertical'.
    layout = plan.get('layout_used')
    if not layout:
        if plan.get('header_row_used'):
            layout = 'horizontal_rows'
        elif column_layout.get('is_column_format'):
            layout = 'column_per_product'
        else:
            layout = 'vertical'
    elif layout == 'vertical' and plan.get('header_row_used') and labels_found_count >= 5:
        # A real horizontal header row was found — never vertical-dump over it.
        layout = 'horizontal_rows'
    plan['layout_used'] = layout

    print(
        f"[FILL] resolved_layout={layout!r} sheet={plan['sheet_used']!r} "
        f"header_row={plan.get('header_row_used')} "
        f"first_data_row={plan.get('first_data_row_used')} "
        f"labels={labels_found_count} products={len(products)} fill_mode={fill_mode!r}",
        flush=True,
    )

    if layout == 'horizontal_rows':
        # Single authoritative pass — precomputed Step 3 values applied INSIDE,
        # with coordinates rebuilt for the horizontal layout (no vertical dump).
        print("[FILL] -> fill_horizontal_rows ONLY (no other write path runs)", flush=True)
        total_filled, all_issues, labels_found_count = fill_horizontal_rows(
            ws, products, plan, precomputed,
        )
        return total_filled, all_issues, labels_found_count

    if layout == 'horizontal_columns':
        print("[FILL] -> fill_horizontal_columns ONLY", flush=True)
        spec = {
            'label_column': plan['label_col'],
            'first_data_column': plan.get('first_data_column') or plan['value_col'],
            'value_column': plan['value_col'],
            'example_rows': plan.get('example_rows'),
        }
        total_filled, all_issues = fill_horizontal_columns(ws, products, spec, plan['label_col'])
        return total_filled, all_issues, labels_found_count

    if layout == 'column_per_product' or column_layout.get('is_column_format'):
        print("[FILL] -> column_per_product fill (PRODUCT N columns)", flush=True)
        label_col = column_layout['label_col']
        template_col = column_layout['product_col_start']
        data_start_row = (column_layout['header_row'] or 1) + 2
        for i, product in enumerate(products):
            col = template_col + i
            add_product_column_headers(ws, column_layout, i, col)
            if i > 0:
                copy_data_cell_format(ws, column_layout, template_col, col, data_start_row)
            filled, issues = fill_single_sheet(ws, product, value_col=col, label_col=label_col)
            total_filled += filled
            all_issues.extend(issues)
        return total_filled, all_issues, labels_found_count

    # ─────────────────────────────────────────────────────────────────────
    # VERTICAL ONLY beyond this point. This is the ONLY code that writes
    # down a single value column (the legacy Suma-style fill). It must never
    # run for a horizontal form.
    # ─────────────────────────────────────────────────────────────────────
    if layout != 'vertical':
        print(
            f"[FILL] WARNING: unexpected layout {layout!r} — refusing to run vertical "
            f"fill; treating as no-op to avoid column dump.",
            flush=True,
        )
        return total_filled, all_issues, labels_found_count

    use_tabs = fill_mode == 'tabs' or (fill_mode == 'auto' and len(products) > 1)
    val_col = plan['value_col']
    label_col = plan['label_col']
    example_rows = plan.get('example_rows') or []
    template_sheet_name = plan['sheet_used']
    other_sheets = plan.get('other_sheets') or set()
    print(
        f"[FILL] -> VERTICAL fill (use_tabs={use_tabs}, value_col={val_col}, "
        f"label_col={label_col})",
        flush=True,
    )

    if use_tabs and len(products) > 1:
        print(f"[TABS_FILL] assigning {len(products)} products to tabs:", flush=True)
        for i, p in enumerate(products):
            print(
                f"[TABS_FILL] index={i} tab=Product {i + 1} "
                f"sku={safe_str(p.get('sku_code'))!r} name={safe_str(p.get('product_name'))!r}",
                flush=True,
            )
        for sheet_name in list(wb.sheetnames):
            if sheet_name == template_sheet_name or sheet_name in other_sheets:
                continue
            if re.match(r'^product\s*\d+$', sheet_name.strip(), re.IGNORECASE):
                del wb[sheet_name]
        clear_vertical_value_column(
            wb[template_sheet_name], val_col, example_rows,
        )
        filled, issues = fill_single_sheet(
            wb[template_sheet_name], products[0], val_col, label_col, example_rows,
            precomputed_by_label=pre_by_index.get(0, {}),
            product_index=0, debug_tab_fill=True,
        )
        total_filled += filled
        all_issues.extend(issues)
        if (re.match(r'^product\s*\d+$', template_sheet_name.strip(), re.IGNORECASE) and
                not re.match(r'^product\s*1$', template_sheet_name.strip(), re.IGNORECASE)):
            wb[template_sheet_name].title = 'Product 1'
            template_sheet_name = 'Product 1'
        template_idx = wb.sheetnames.index(template_sheet_name)
        for i, product in enumerate(products[1:], start=2):
            new_ws = wb.copy_worksheet(wb[template_sheet_name])
            new_ws.title = f'Product {i}'
            for dv in wb[template_sheet_name].data_validations.dataValidation:
                new_ws.add_data_validation(deepcopy(dv))
            current_idx = wb.sheetnames.index(new_ws.title)
            wb.move_sheet(new_ws.title, offset=(template_idx + i - 1) - current_idx)
            clear_vertical_value_column(new_ws, val_col, example_rows)
            filled, issues = fill_single_sheet(
                new_ws, product, val_col, label_col, example_rows,
                precomputed_by_label=pre_by_index.get(i - 1, {}),
                product_index=i - 1, debug_tab_fill=True,
            )
            total_filled += filled
            all_issues.extend(issues)
        plan['tabs_fill_used'] = True
        return total_filled, all_issues, labels_found_count

    if fill_mode == 'columns':
        for i, product in enumerate(products):
            filled, issues = fill_single_sheet(
                ws, product, value_col=val_col + i, label_col=label_col,
                precomputed_by_label=pre_by_index.get(i, {}),
            )
            total_filled += filled
            all_issues.extend(issues)
        return total_filled, all_issues, labels_found_count

    filled, issues = fill_single_sheet(
        ws, products[0], val_col, label_col, example_rows,
        precomputed_by_label=pre_by_index.get(0, {}),
    )
    total_filled += filled
    all_issues.extend(issues)
    return total_filled, all_issues, labels_found_count


def _fill_product_row(ws, row_num, headers, product):
    filled = 0
    unresolved = []
    for col, label in headers.items():
        existing = ws.cell(row=row_num, column=col)
        if is_formula_cell(existing):
            continue
        value = map_field(label, product)
        if value is not None:
            if value in ('None', 'null', ''):
                value = 'N/A'
            if safe_write(ws, row_num, col, value):
                filled += 1
        else:
            unresolved.append((col, label))
    if unresolved:
        resolved = claude_resolve_fields([l for _, l in unresolved], product)
        for col, label in unresolved:
            if label in resolved and resolved[label]:
                existing = ws.cell(row=row_num, column=col)
                if not is_formula_cell(existing):
                    val = resolved[label]
                    if val in ('None', 'null', ''):
                        val = 'N/A'
                    if safe_write(ws, row_num, col, val):
                        filled += 1
    return filled


def fill_horizontal_columns(ws, products, form_spec, label_col=None):
    print(f"[FILL_FN] fill_horizontal_columns ENTER sheet={ws.title!r}", flush=True)
    label_col = label_col or form_spec.get('label_column') or 2
    base_col = form_spec.get('first_data_column') or form_spec.get('value_column') or 3
    total_filled = 0
    all_issues = []
    for i, product in enumerate(products):
        filled, issues = fill_single_sheet(
            ws, product, value_col=base_col + i, label_col=label_col,
            example_rows=form_spec.get('example_rows'),
        )
        total_filled += filled
        all_issues.extend(issues)
    return total_filled, all_issues


def fill_using_form_spec(wb, products, form_spec, fill_mode='auto'):
    plan = resolve_fill_plan(wb, form_spec, fill_mode)
    total_filled, all_issues, _ = execute_fill(wb, products, plan, fill_mode)
    return total_filled, all_issues


def add_product_column_headers(ws, layout, product_index, product_col):
    header_row = layout['header_row']
    template_col = layout['product_col_start']
    product_num = product_index + 1

    if not header_row:
        return

    for r_offset in range(2):
        r = header_row + r_offset
        template_cell = ws.cell(row=r, column=template_col)
        new_cell = ws.cell(row=r, column=product_col)

        if r_offset == 0:
            new_cell.value = f"PRODUCT {product_num}"
        else:
            new_cell.value = f"Fill in product {product_num} details below ↓"

        if not isinstance(template_cell, MergedCell):
            if template_cell.fill and template_cell.fill.fill_type == 'solid':
                new_cell.fill = copy(template_cell.fill)
            if template_cell.font:
                new_cell.font = copy(template_cell.font)
            if template_cell.alignment:
                new_cell.alignment = copy(template_cell.alignment)


def copy_data_cell_format(ws, layout, template_col, new_col, data_start_row):
    for row in ws.iter_rows(min_row=data_start_row):
        tc = ws.cell(row=row[0].row, column=template_col)
        nc = ws.cell(row=row[0].row, column=new_col)
        if isinstance(tc, MergedCell) or isinstance(nc, MergedCell):
            continue
        if tc.fill and tc.fill.fill_type == 'solid':
            nc.fill = copy(tc.fill)


@app.post("/fill")
async def fill_nlf(req: FillRequest):
    try:
        if not req.products:
            raise HTTPException(status_code=422, detail="No products provided")

        file_bytes = base64.b64decode(req.file_base64)
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
        _FillGuard.reset()
        _WriteTracker.reset()
        formula_count_before = count_formula_cells(wb)

        total_filled = 0
        all_issues_fixed = []

        # Resolve the layout ONCE — this is the single authoritative layout.
        plan = resolve_fill_plan(wb, req.form_spec, req.fill_mode)
        resolved_layout = plan['layout_used']

        # EXACTLY ONE fill pass. For horizontal_rows the precomputed Step 3 values
        # are applied inside the pass with horizontal coordinates. For other layouts
        # the precomputed values (which carry layout-valid coordinates) overlay the
        # map_field fill afterwards — never as a conflicting independent layout.
        total_filled, all_issues_fixed, labels_found_count = execute_fill(
            wb, req.products, plan, req.fill_mode,
            precomputed=req.precomputed_mappings,
        )

        if resolved_layout != 'horizontal_rows' and req.precomputed_mappings and not plan.get('tabs_fill_used'):
            precomputed_filled = apply_precomputed_mappings(
                wb, req.precomputed_mappings, plan,
                fill_mode=req.fill_mode,
                num_products=len(req.products),
            )
            labels_found_count = max(labels_found_count, len(req.precomputed_mappings))
            total_filled = max(total_filled, precomputed_filled)

        # FIX 4 — conflict guard: never return a file corrupted by two layout passes.
        conflict = check_write_conflict(plan, len(req.products))
        if conflict:
            raise HTTPException(status_code=422, detail=conflict)

        formula_count_after = count_formula_cells(wb)
        if _FillGuard.formula_blocks > 0:
            raise HTTPException(
                status_code=500,
                detail=(
                    f'Formula protection violation: blocked {_FillGuard.formula_blocks} '
                    f'write attempt(s) to formula cells. File not returned.'
                ),
            )
        if formula_count_after < formula_count_before:
            raise HTTPException(
                status_code=500,
                detail=(
                    f'Formula protection violation: formula cell count decreased '
                    f'from {formula_count_before} to {formula_count_after}. '
                    f'File not returned to prevent damage.'
                ),
            )

        if total_filled == 0:
            raise HTTPException(
                status_code=422,
                detail=(
                    f"No fields were filled. Layout detected: {plan.get('layout_used')}. "
                    f"Data sheet: {plan.get('sheet_used')}. "
                    f"Header row: {plan.get('header_row_used')}. "
                    f"First data row: {plan.get('first_data_row_used')}. "
                    f"Labels found: {labels_found_count}. "
                    f"Products received: {len(req.products)}. "
                    f"Precomputed mappings received: {len(req.precomputed_mappings or [])}. "
                    f"This indicates a layout detection or field matching failure."
                ),
            )

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        date_str = datetime.date.today().isoformat()
        name = req.products[0].get('product_name', 'Product') if req.products else 'Product'
        if len(req.products) > 1:
            filename = f"{req.retailer_name}_NLF_{len(req.products)}_products_{date_str}.xlsx".replace(" ", "_")
        else:
            filename = f"{req.retailer_name}_{name}_{date_str}.xlsx".replace(" ", "_")

        return {
            "file_base64": base64.b64encode(output.read()).decode(),
            "filename": filename,
            "fields_filled": total_filled,
            "products_filled": len(req.products),
            "fill_mode": req.fill_mode,
            "verification_issues_fixed": len(all_issues_fixed),
            "anthropic_key_present": bool(get_anthropic_api_key()),
        }

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        raise HTTPException(status_code=500, detail=f"{str(e)}\n{traceback.format_exc()}")


@app.get("/health")
async def health():
    return {"status": "ok"}


@app.get("/health/config")
async def health_config():
    """Safe diagnostics — booleans only, never secret values."""
    return anthropic_config_status()


if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)
