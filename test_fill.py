"""Local smoke test for the /fill logic (tabs and columns modes)."""
import asyncio
import base64
import io

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from main import FieldMapping, FillRequest, ProductFill, fill_nlf

GREEN = "FFEAF1DD"


def build_template() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Product 1"
    labels = ["Product Name", "Brand", "RRP", "Vegan"]
    fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
    for i, label in enumerate(labels, start=4):
        ws.cell(row=i, column=2).value = label
        ws.cell(row=i, column=3).fill = fill
    dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    dv.add("C7")
    ws.add_data_validation(dv)
    wb.create_sheet("Additional")
    wb.create_sheet("C.O.O List")
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def product(name: str, rrp: str) -> ProductFill:
    return ProductFill(
        product_name=name,
        mappings=[
            FieldMapping(field_name="Product Name", mapped_value=name, status="filled", row=4, col="C"),
            FieldMapping(field_name="Brand", mapped_value="Nine Streets", status="filled"),
            FieldMapping(field_name="RRP", mapped_value=rrp, status="filled"),
            FieldMapping(field_name="Vegan", mapped_value="Yes", status="filled"),
            FieldMapping(field_name="Unknown Field", mapped_value=None, status="missing"),
        ],
    )


def run(mode: str, products):
    req = FillRequest(
        file_base64=base64.b64encode(build_template()).decode(),
        fill_mode=mode,
        products=products,
        retailer_name="Suma",
    )
    res = asyncio.run(fill_nlf(req))
    wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(res["file_base64"])))
    return res, wb


# --- tabs mode, 3 products ---
res, wb = run("tabs", [product("Granola A", "4.99"), product("Granola B", "5.49"), product("Granola C", "5.99")])
assert res["fields_filled"] == 12, res
assert wb.sheetnames == ["Product 1", "Product 2", "Product 3", "Additional", "C.O.O List"], wb.sheetnames
for sheet, name, rrp in [("Product 1", "Granola A", "4.99"), ("Product 2", "Granola B", "5.49"), ("Product 3", "Granola C", "5.99")]:
    ws = wb[sheet]
    assert ws["C4"].value == name, (sheet, ws["C4"].value)
    assert ws["C6"].value == rrp
    assert ws["C4"].fill.start_color.rgb == GREEN, (sheet, ws["C4"].fill.start_color.rgb)
    dvs = ws.data_validations.dataValidation
    assert len(dvs) == 1 and dvs[0].formula1 == '"Yes,No"', (sheet, dvs)
print("tabs mode OK — sheets:", wb.sheetnames)

# --- columns mode, 3 products ---
res, wb = run("columns", [product("Granola A", "4.99"), product("Granola B", "5.49"), product("Granola C", "5.99")])
assert res["fields_filled"] == 12, res
ws = wb["Product 1"]
assert ws["C4"].value == "Granola A" and ws["D4"].value == "Granola B" and ws["E4"].value == "Granola C"
assert ws["C6"].value == "4.99" and ws["D6"].value == "5.49" and ws["E6"].value == "5.99"
assert wb.sheetnames == ["Product 1", "Additional", "C.O.O List"]
print("columns mode OK — C/D/E filled")

# --- legacy single-product payload ---
req = FillRequest(
    file_base64=base64.b64encode(build_template()).decode(),
    mappings=product("Solo", "3.99").mappings,
    product_name="Solo",
    retailer_name="Suma",
)
res = asyncio.run(fill_nlf(req))
wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(res["file_base64"])))
assert wb["Product 1"]["C4"].value == "Solo"
assert res["filename"].startswith("Suma_Solo_")
print("legacy single-product OK —", res["filename"])
print("ALL TESTS PASSED")
